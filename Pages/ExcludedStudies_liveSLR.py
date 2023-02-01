import os
import time

import openpyxl
import pandas as pd
from selenium.webdriver.support.wait import WebDriverWait

from Pages.Base import Base
from Pages.OpenLiveSLRPage import LiveSLRPage
from Pages.SLRReportPage import SLRReport
from utilities.customLogger import LogGen
from utilities.logScreenshot import cLogScreenshot


class ExcludedStudies_liveSLR(Base):

    """Constructor of the ExcludedStudies_liveSLR class"""
    def __init__(self, driver, extra):
        # initializing the driver from base class
        super().__init__(driver, extra)  
        self.extra = extra
        # Instantiate the Base class
        self.base = Base(self.driver, self.extra)
        # Creating object of liveslrpage class
        self.liveslrpage = LiveSLRPage(self.driver, self.extra)
        # Creating object of slrreport class
        self.slrreport = SLRReport(self.driver, extra)
        # Instantiate the logger class
        self.logger = LogGen.loggen()
        # Instantiate the logScreenshot class
        self.LogScreenshot = cLogScreenshot(self.driver, self.extra)
        # Instantiate webdriver wait class
        self.wait = WebDriverWait(driver, 20)
    
    def get_source_template(self, filepath):
        file = pd.read_excel(filepath)
        expectedfilepath = list(os.getcwd()+file['ExpectedSourceTemplateFile'].dropna())
        return expectedfilepath

    def get_population_data(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        pop = df.loc[df['Name'] == locatorname]['Population'].dropna().to_list()
        pop_button = df.loc[df['Name'] == locatorname]['Population_Radio_button'].dropna().to_list()
        result = [[pop[i], pop_button[i]] for i in range(0, len(pop))]
        return result
    
    def get_slrtype_data(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        slrtype = df.loc[df['Name'] == locatorname]['slrtype'].dropna().to_list()
        slrtype_button = df.loc[df['Name'] == locatorname]['slrtype_Radio_button'].dropna().to_list()
        result = [[slrtype[i], slrtype_button[i]] for i in range(0, len(slrtype))]
        return result

    def get_additional_criteria_data(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        criteria = df.loc[df['Name'] == locatorname]['AddtionalParam'].dropna().to_list()
        criteria_btn = df.loc[df['Name'] == locatorname]['AddtionalParam_button'].dropna().to_list()
        section_name = df.loc[df['Name'] == locatorname]['sectionname'].dropna().to_list()
        result = [[criteria[i], criteria_btn[i], section_name[i]] for i in range(0, len(criteria))]
        return result
    
    def presenceof_excludedstudiesliveslr_into_excelreport(self, locatorname, filepath, env):
        pop_data = self.get_population_data(filepath, locatorname)

        slr_type = self.get_slrtype_data(filepath, locatorname)

        add_criteria = self.get_additional_criteria_data(filepath, locatorname)

        # Go to live slr page
        self.liveslrpage.go_to_liveslr("SLR_Homepage", env)
        time.sleep(2)
        self.slrreport.select_data(f"{pop_data[0][0]}", f"{pop_data[0][1]}", env)
        self.slrreport.select_data(f"{slr_type[0][0]}", f"{slr_type[0][1]}", env)
        self.slrreport.select_sub_section(f"{add_criteria[0][0]}", f"{add_criteria[0][1]}", env, f"{add_criteria[0][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[1][0]}", f"{add_criteria[1][1]}", env, f"{add_criteria[1][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[2][0]}", f"{add_criteria[2][1]}", env, f"{add_criteria[2][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[3][0]}", f"{add_criteria[3][1]}", env, f"{add_criteria[3][2]}")

        self.slrreport.generate_download_report("excel_report", env)
        # time.sleep(5)
        # excel_filename = self.slrreport.getFilenameAndValidate(180)
        # excel_filename = self.slrreport.get_latest_filename(UnivWaitFor=180)
        excel_filename = self.slrreport.get_and_validate_filename(filepath)

        self.LogScreenshot.fLogScreenshot(message=f"*****Check Presence of Excluded studies - LiveSLR Tab in Complete "
                                                  f"Excel Report*****",
                                          pass_=True, log=True, screenshot=False)
        excel_data = openpyxl.load_workbook(f'ActualOutputs//{excel_filename}')
        if 'Excluded studies - LiveSLR' in excel_data.sheetnames:
            self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' tab is present in Downloaded "
                                                      f"Complete Excel Report",
                                              pass_=True, log=True, screenshot=False)
        else:
            self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' tab is missing in "
                                                      f"Complete Excel Report",
                                              pass_=False, log=True, screenshot=False)
            raise Exception("'Excluded studies - LiveSLR' tab is missing in Complete Excel Report")

        excel_sheet = excel_data[f'Excluded studies - LiveSLR']
        if excel_sheet['J1'].value == 'Back To Toc':
            self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is present in 'Excluded studies - LiveSLR'"
                                                      f" sheet",
                                              pass_=True, log=True, screenshot=False)
        else:
            self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is not present in 'Excluded studies - "
                                                      f"LiveSLR' sheet",
                                              pass_=False, log=True, screenshot=False)
            raise Exception(f"'Back To Toc' option is not present in 'Excluded studies - LiveSLR' sheet")
        
        toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
        col_data = list(toc_sheet.iloc[:, 1])
        if f'Excluded studies - LiveSLR' in col_data:
            self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' is present in TOC sheet.",
                                              pass_=True, log=True, screenshot=False)
        else:
            self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' is not present in TOC sheet. "
                                                      f"Available Data from TOC sheet: {col_data}",
                                              pass_=False, log=True, screenshot=False)
            raise Exception("'Excluded studies - LiveSLR' is not present in TOC sheet.")
        
    def presenceof_columnnames_in_excludedstudiesliveslrtab(self, locatorname, filepath, index, env):
        pop_data = self.get_population_data(filepath, locatorname)

        slr_type = self.get_slrtype_data(filepath, locatorname)

        add_criteria = self.get_additional_criteria_data(filepath, locatorname)

        source_template = self.get_source_template(filepath)

        # Go to live slr page
        self.liveslrpage.go_to_liveslr("SLR_Homepage", env)
        time.sleep(2)
        self.slrreport.select_data(f"{pop_data[0][0]}", f"{pop_data[0][1]}", env)
        self.slrreport.select_data(f"{slr_type[0][0]}", f"{slr_type[0][1]}", env)
        self.slrreport.select_sub_section(f"{add_criteria[0][0]}", f"{add_criteria[0][1]}", env, f"{add_criteria[0][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[1][0]}", f"{add_criteria[1][1]}", env, f"{add_criteria[1][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[2][0]}", f"{add_criteria[2][1]}", env, f"{add_criteria[2][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[3][0]}", f"{add_criteria[3][1]}", env, f"{add_criteria[3][2]}")

        self.slrreport.generate_download_report("excel_report", env)
        # time.sleep(5)
        # excel_filename = self.slrreport.getFilenameAndValidate(180)
        # excel_filename = self.slrreport.get_latest_filename(UnivWaitFor=180)
        excel_filename = self.slrreport.get_and_validate_filename(filepath)

        self.LogScreenshot.fLogScreenshot(message=f"*****Check Presence of expected columns in Excluded studies - "
                                                  f"LiveSLR Tab*****",
                                          pass_=True, log=True, screenshot=False)
        source_template_sheet = openpyxl.load_workbook(f'{source_template[0]}')
        sheets = source_template_sheet.sheetnames

        sourcefile = pd.read_excel(f'{source_template[0]}', sheet_name=sheets[index])
        actualexcel = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name='Excluded studies - LiveSLR',
                                    skiprows=2)
        
        source_col_val = [item for item in list(sourcefile.columns.values) if str(item) != 'nan']
        compex_col_val = [item for item in list(actualexcel.columns.values) if str(item) != 'nan']
        # Removing the unnamed column from the downloaded report. This column got created for 'Back To TOC' option
        compex_col_val.pop(-1)

        if source_col_val == compex_col_val:
            self.LogScreenshot.fLogScreenshot(message=f"Expected column names are present in Complete Excel Report. "
                                                      f"Column names are: {compex_col_val}",
                                              pass_=True, log=True, screenshot=False)
        else:
            self.LogScreenshot.fLogScreenshot(message=f"Expected column names are not present in Complete Excel Report."
                                                      f"\nSource Excel Column names are : {source_col_val},"
                                                      f"\nComplete Excel Column names are : {compex_col_val}",
                                              pass_=False, log=True, screenshot=False)
            raise Exception("Expected column names are not present in Complete Excel Report")

    def validate_excludedstudiesliveslrtab_and_contents_into_excelreport(self, locatorname, filepath, index, env):
        pop_data = self.get_population_data(filepath, locatorname)

        slr_type = self.get_slrtype_data(filepath, locatorname)

        add_criteria = self.get_additional_criteria_data(filepath, locatorname)

        source_template = self.get_source_template(filepath)

        # Go to live slr page
        self.liveslrpage.go_to_liveslr("SLR_Homepage", env)
        time.sleep(2)
        self.slrreport.select_data(f"{pop_data[0][0]}", f"{pop_data[0][1]}", env)
        self.slrreport.select_data(f"{slr_type[0][0]}", f"{slr_type[0][1]}", env)
        self.slrreport.select_sub_section(f"{add_criteria[0][0]}", f"{add_criteria[0][1]}", env, f"{add_criteria[0][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[1][0]}", f"{add_criteria[1][1]}", env, f"{add_criteria[1][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[2][0]}", f"{add_criteria[2][1]}", env, f"{add_criteria[2][2]}")
        self.slrreport.select_sub_section(f"{add_criteria[3][0]}", f"{add_criteria[3][1]}", env, f"{add_criteria[3][2]}")

        self.slrreport.generate_download_report("excel_report", env)
        # time.sleep(5)
        # excel_filename = self.slrreport.getFilenameAndValidate(180)
        # excel_filename = self.slrreport.get_latest_filename(UnivWaitFor=180)
        excel_filename = self.slrreport.get_and_validate_filename(filepath)

        self.LogScreenshot.fLogScreenshot(message=f"*****Check Presence of Excluded studies - LiveSLR Tab in Complete "
                                                  f"Excel Report*****",
                                          pass_=True, log=True, screenshot=False)
        excel_data = openpyxl.load_workbook(f'ActualOutputs//{excel_filename}')
        if 'Excluded studies - LiveSLR' in excel_data.sheetnames:
            self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' tab is present in Downloaded "
                                                      f"Complete Excel Report",
                                              pass_=True, log=True, screenshot=False)

            excel_sheet = excel_data[f'Excluded studies - LiveSLR']
            if excel_sheet['J1'].value == 'Back To Toc':
                self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is present in 'Excluded studies - "
                                                          f"LiveSLR' sheet",
                                                  pass_=True, log=True, screenshot=False)
            else:
                self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is not present in 'Excluded studies - "
                                                          f"LiveSLR' sheet",
                                                  pass_=False, log=True, screenshot=False)
                raise Exception(f"'Back To Toc' option is not present in 'Excluded studies - LiveSLR' sheet")
            
            toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
            col_data = list(toc_sheet.iloc[:, 1])
            if f'Excluded studies - LiveSLR' in col_data:
                self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' is present in TOC sheet.",
                                                  pass_=True, log=True, screenshot=False)
            else:
                self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' is not present in TOC sheet. "
                                                          f"Available Data from TOC sheet: {col_data}",
                                                  pass_=False, log=True, screenshot=False)
                raise Exception("'Excluded studies - LiveSLR' is not present in TOC sheet.")
        
            # Excluded studies - LiveSLR sheet comparison with Expected results
            self.LogScreenshot.fLogScreenshot(message=f"*****Excluded studies - LiveSLR Sheet Content Comparison "
                                                      f"Started*****",
                                              pass_=True, log=True, screenshot=False)
            source_template_sheet = openpyxl.load_workbook(f'{source_template[0]}')
            sheets = source_template_sheet.sheetnames

            sourcefile = pd.read_excel(f'{source_template[0]}', sheet_name=sheets[index])
            actualexcel = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name='Excluded studies - LiveSLR',
                                        skiprows=2)
            
            cols = list(sourcefile.columns.values)

            # Check the length of 1st column from the report to make sure number of rows are as expected
            source_col_len = sourcefile["Study Identifier"]
            actual_col_len = actualexcel["Study Identifier"]

            source_col_len = [item for item in source_col_len if str(item) != 'nan']
            actual_col_len = [item for item in actual_col_len if str(item) != 'nan']

            if len(source_col_len) == len(actual_col_len):
                self.LogScreenshot.fLogScreenshot(message=f"Elements length is matching between Source Template and "
                                                          f"Complete Excel Report.\n"
                                                          f"Source Elements Length: {len(source_col_len)}\n"
                                                          f"Excel Elements Length: {len(actual_col_len)}\n",
                                                  pass_=True, log=True, screenshot=False)

                # Content comparison between Source file and Complete Excel report
                for col in cols:
                    source_col = sourcefile[col]
                    actual_col = actualexcel[col]

                    source_col = [item for item in source_col if str(item) != 'nan']
                    actual_col = [item for item in actual_col if str(item) != 'nan']

                    comparison_result = self.slrreport.list_comparison_between_reports_data(source_col, actual_col)

                    if len(comparison_result) == 0:
                        self.LogScreenshot.fLogScreenshot(message=f"Contents in column '{col}' are matching between "
                                                                  f"Source Template and Complete Excel Report",
                                                          pass_=True, log=True, screenshot=False)
                    else:
                        self.LogScreenshot.fLogScreenshot(message=f"Contents in column '{col}' are not matching. "
                                                                  f"Duplicate values are arranged in following order ->"
                                                                  f" Source File, Complete Excel. {comparison_result}",
                                                          pass_=False, log=True, screenshot=False)
                        raise Exception("Contents are not matching between Source template and Complete Excel report")
            else:
                self.LogScreenshot.fLogScreenshot(message=f"Elements length is not matching between Source Template "
                                                          f"and Complete Excel Report. "
                                                          f"Source Template Elements Length: {len(source_col_len)}\n"
                                                          f"Excel Elements Length: {len(actual_col_len)}\n",
                                                  pass_=False, log=True, screenshot=False)
                raise Exception(f"Elements length is not matching between Source Template and Complete Excel Report.")

            self.LogScreenshot.fLogScreenshot(message=f"*****Excluded studies - LiveSLR Sheet Content Comparison "
                                                      f"Completed*****",
                                              pass_=True, log=True, screenshot=False)
        else:
            self.LogScreenshot.fLogScreenshot(message=f"'Excluded studies - LiveSLR' tab is missing in Complete "
                                                      f"Excel Report",
                                              pass_=False, log=True, screenshot=False)
            raise Exception("'Excluded studies - LiveSLR' tab is missing in Complete Excel Report")
