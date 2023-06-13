import os
from pathlib import Path
import time
import openpyxl
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from Pages.Base import Base
from Pages.ExtendedBasePage import ExtendedBase
from Pages.ManagePopulationsPage import ManagePopulationsPage
from Pages.OpenLiveSLRPage import LiveSLRPage
from Pages.SLRReportPage import SLRReport
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from utilities.logScreenshot import cLogScreenshot
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import Select


class ManageQADataPage(Base):

    """Constructor of the ManageQAData Page class"""
    def __init__(self, driver, extra):
        # initializing the driver from base class
        super().__init__(driver, extra)  
        self.extra = extra
        # Instantiate the Base class
        self.base = Base(self.driver, self.extra)
        # Creating object of ExtendedBase class
        self.exbase = ExtendedBase(self.driver, extra)        
        # Instantiate the logger class
        self.logger = LogGen.loggen()
        # Instantiate the logScreenshot class
        self.LogScreenshot = cLogScreenshot(self.driver, self.extra)
        # Creating object of liveslrpage class
        self.liveslrpage = LiveSLRPage(self.driver, extra)
        # Creating object of slrreport class
        self.slrreport = SLRReport(self.driver, extra)
        # Creating object of ManagePopulationsPage class
        self.mngpoppage = ManagePopulationsPage(self.driver, extra)
        # Instantiate webdriver wait class
        self.wait = WebDriverWait(driver, 10)
    
    def presence_of_elements(self, locator, env):
        self.wait.until(ec.presence_of_element_located((getattr(By, self.locatortype(locator, env)),
                                                        self.locatorpath(locator, env))))
    
    # Reading Population data for ManageQAData Page
    def get_manageqa_pop_data(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        pop = df.loc[df['Name'] == locatorname]['QA_Population'].dropna().to_list()
        return pop

    def get_qa_file_details(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        study_type = df.loc[df['Name'] == locatorname]['Study_Types'].dropna().to_list()
        qa_file = df.loc[df['Name'] == locatorname]['QA_Excel_Files'].dropna().to_list()
        result = [[study_type[i], os.getcwd()+qa_file[i]] for i in range(0, len(study_type))]
        return result
    
    def get_qa_file_details_override(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        study_type = df.loc[df['Name'] == locatorname]['Study_Types'].dropna().to_list()
        qa_file = df.loc[df['Name'] == locatorname]['Override_QA_Excel_Files'].dropna().to_list()
        result = [[study_type[i], os.getcwd()+qa_file[i]] for i in range(0, len(study_type))]
        return result
    
    def get_invalid_qa_file_details(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        study_type = df.loc[df['Name'] == locatorname]['Study_Types'].dropna().to_list()
        qa_file = df.loc[df['Name'] == locatorname]['Invalid_Files'].dropna().to_list()
        result = [[study_type[i], os.getcwd()+qa_file[i]] for i in range(0, len(study_type))]
        return result

    def access_manageqadata_page_elements(self, locatorname, filepath, env):

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
		# Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["QA_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        self.click("select_pop_dropdown", env)
                        self.LogScreenshot.fLogScreenshot(message=f"Population dropdown is accessible. Listed elements are:",
                                                  pass_=True, log=True, screenshot=True)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)
                        
                        self.click("select_stdy_type_dropdown", env)
                        self.LogScreenshot.fLogScreenshot(message=f"SLR Type dropdown is accessible. Listed elements are:",
                                                  pass_=True, log=True, screenshot=True)
                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.input_text("qa_checklist_name", f"QAName_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_citation", f"QACitation_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_reference", f"QAReference_{i[0]}_{k[0]}", env)
                        self.input_text("qa_excel_file_upload", k[1], env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(message=f'User is able to enter the details for Population : '
                                                                f'{i[0]} -> SLR Type : {selected_slr_val}',
                                                        pass_=True, log=True, screenshot=True)
                        
                        self.presence_of_elements("upload_save_button", env)
                        self.presence_of_elements("delete_file_button", env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(message=f"ManageQAData Page elements are accessible for "
                                                                f"Population : {pop_value} -> SLR Type : {selected_slr_val}.",
                                                        pass_=True, log=True, screenshot=True)
        except Exception:
            raise Exception("Unable to access the Manage QA Data page elements")

    def add_manage_qa_data_with_invalidfile(self, locatorname, filepath, env):
        expected_error_text = "Quality assessment data could not be updated due to validation errors"

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
		# Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["Invalid_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)
                        
                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.input_text("qa_checklist_name", f"QAName_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_citation", f"QACitation_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_reference", f"QAReference_{i[0]}_{k[0]}", env)
                        self.input_text("qa_excel_file_upload", k[1], env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(message=f'User is able to enter the details for Population : '
                                                                f'{i[0]} -> SLR Type : {selected_slr_val}',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("upload_save_button", env)
                        time.sleep(3)
                        # actual_upload_status_text = self.get_text("get_status_text", env, UnivWaitFor=10)
                        actual_error_text = self.get_status_text("get_status_text", env)
                        # time.sleep(2)

                        if actual_error_text == expected_error_text:
                            self.LogScreenshot.fLogScreenshot(message=f"File with invalid format is not uploaded as expected. "
                                                                    f"Invalid file is '{Path(f'{k[1]}').name}'",
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f"Unable to find status message while uploading File "
                                                                    f"with invalid format for Population : {pop_value} -> "
                                                                    f"SLR Type : {selected_slr_val}. Invalid file is "
                                                                    f"'{Path(f'{k[1]}').name}'. Actual status message is "
                                                                    f"{actual_error_text} and Expected status message is "
                                                                    f"{expected_error_text}",
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while uploading File with invalid format "
                                            f"for Population : {pop_value} -> SLR Type : {selected_slr_val}. "
                                            f"Invalid file is '{Path(f'{i[1]}').name}'")
        except Exception:
            raise Exception("Unable to upload the Manage QA Data")

    def add_multiple_manage_qa_data(self, locatorname, filepath, env):
        expected_upload_status_text = 'QA File successfully uploaded'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
		# Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["QA_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)
                        
                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.input_text("qa_checklist_name", f"QAName_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_citation", f"QACitation_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_reference", f"QAReference_{i[0]}_{k[0]}", env)
                        self.input_text("qa_excel_file_upload", k[1], env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(message=f'User is able to enter the details for Population : '
                                                                f'{i[0]} -> SLR Type : {selected_slr_val}',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("upload_save_button", env)
                        time.sleep(2)
                        # actual_upload_status_text = self.get_text("get_status_text", env, UnivWaitFor=10)
                        actual_upload_status_text = self.get_status_text("get_status_text", env)
                        # time.sleep(2)

                        if actual_upload_status_text == expected_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'QA File upload is success for Population : '
                                                                    f'{pop_value} -> SLR Type : {selected_slr_val}.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f'Unable to find status message while uploading QA File '
                                                                    f'for Population : {pop_value} -> SLR Type : {selected_slr_val}. '
                                                                    f'Actual status message is {actual_upload_status_text} '
                                                                    f'and Expected status message is '
                                                                    f'{expected_upload_status_text}',
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while uploading QA File for Population : "
                                            f"{pop_value} -> SLR Type : {selected_slr_val}.")
        except Exception:
            raise Exception("Unable to upload the Manage QA Data")

    def overwrite_multiple_manage_qa_data(self, locatorname, filepath, env):
        expected_upload_status_text = 'QA File successfully uploaded'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
		# Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["Override_QA_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)
                        
                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.input_text("qa_checklist_name", f"QAName_{i[0]}_{k[0]}_Override", env)
                        self.input_text("qa_checklist_citation", f"QACitation_{i[0]}_{k[0]}_Override", env)
                        self.input_text("qa_checklist_reference", f"QAReference_{i[0]}_{k[0]}_Override", env)
                        self.input_text("qa_excel_file_upload", k[1], env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(message=f'User is able to override the details for Population : '
                                                                f'{pop_value} -> SLR Type : {selected_slr_val}',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("upload_save_button", env)
                        time.sleep(2)
                        # actual_upload_status_text = self.get_text("get_status_text", env, UnivWaitFor=10)
                        actual_upload_status_text = self.get_status_text("get_status_text", env)
                        # time.sleep(2)

                        if actual_upload_status_text == expected_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'Updating the existing QA File is success for '
                                                                    f'Population : {pop_value} -> SLR Type : {selected_slr_val}.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f'Unable to find status message while Updating the '
                                                                    f'existing QA File for Population : {pop_value} -> '
                                                                    f'SLR Type : {selected_slr_val}. Actual status message is '
                                                                    f'{actual_upload_status_text} and Expected status '
                                                                    f'message is {expected_upload_status_text}',
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while Updating the existing QA File for "
                                            f"Population : {pop_value} -> SLR Type : {selected_slr_val}")
        except Exception:
            raise Exception("Unable to overwrite the Manage QA Data")

    def del_multiple_manage_qa_data(self, locatorname, filepath, env):
        expected_delete_status_text = 'QA excel file successfully deleted'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
		# Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["QA_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)
                        
                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)
                        self.LogScreenshot.fLogScreenshot(message=f'Selected Data For Deletion :',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("delete_file_button", env)
                        time.sleep(2)
                        self.click("delete_file_popup", env)
                        time.sleep(2)

                        # actual_delete_status_text = self.get_text("get_status_text", env, UnivWaitFor=10)
                        actual_delete_status_text = self.get_status_text("get_status_text", env)
                        # time.sleep(2)

                        if actual_delete_status_text == expected_delete_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'QA File Deletion is success for Population : '
                                                                    f'{pop_value} -> SLR Type : {selected_slr_val}.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f'Unable to find status message while deleting QA File '
                                                                    f'for Population : {pop_value} -> SLR Type : {selected_slr_val}. '
                                                                    f'Actual status message is {actual_delete_status_text} '
                                                                    f'and Expected status message is '
                                                                    f'{expected_delete_status_text}',
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while deleting QA File for Population : "
                                            f"{pop_value} -> SLR Type : {selected_slr_val}.")
        except Exception:
            raise Exception("Unable to delete the existing QA file")

    def compare_qa_file_with_report(self, locatorname, filepath, env):
        expected_upload_status_text = 'QA File successfully uploaded'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])
        
        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["QA_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.input_text("qa_checklist_name", f"QAName_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_citation", f"QACitation_{i[0]}_{k[0]}", env)
                        self.input_text("qa_checklist_reference", f"QAReference_{i[0]}_{k[0]}", env)
                        self.input_text("qa_excel_file_upload", k[1], env)
                        time.sleep(3)
                        self.LogScreenshot.fLogScreenshot(message=f'User is able to enter the details for Population : '
                                                                f'{i[0]} -> SLR Type : {k[0]}',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("upload_save_button", env)
                        time.sleep(3)
                        actual_upload_status_text = self.get_status_text("get_status_text", env, UnivWaitFor=10)

                        if actual_upload_status_text == expected_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'QA File upload is success for Population : '
                                                                    f'{i[0]} -> SLR Type : {k[0]}.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f'Unable to find status message while uploading QA File '
                                                                    f'for Population : {i[0]} -> SLR Type : {k[0]}.',
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while uploading QA File for Population : "
                                            f"{i[0]} -> SLR Type : {k[0]}.")
                    
                    # Go to live slr page
                    self.go_to_page("SLR_Homepage", env)
                    # if j[0] == 'Quality of life':
                    #     j[0] = 'Quality of Life'
                    self.slrreport.select_data(f"{i[0]}", f"{i[1]}", env)
                    self.slrreport.select_data(f"{j[0]}", f"{j[1]}", env)
                    self.slrreport.generate_download_report("excel_report", env)
                    excel_filename = self.slrreport.get_and_validate_filename(filepath)

                    excel_data = openpyxl.load_workbook(f'ActualOutputs//{excel_filename}')
                    if ['Clinical-Interventional', 'Clinical-RWE'] == stdytype:
                        expected_sheet_names = ['Quality Assessment-Intervtnl', 'Quality Assessment-RWE']
                        for index, sheet in enumerate(expected_sheet_names):
                            if sheet in excel_data.sheetnames:
                                self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' sheet is present in complete "
                                                                        f"excel report",
                                                                pass_=True, log=True, screenshot=False)

                                excel_sheet = excel_data[sheet]
                                if excel_sheet['A1'].value == 'Back To Toc':
                                    self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is present",
                                                                    pass_=True, log=True, screenshot=False)
                                else:
                                    self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is not present",
                                                                    pass_=False, log=True, screenshot=False)
                                    raise Exception(f"'Back To Toc' option is not present")
                                
                                toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                                col_data = list(toc_sheet.iloc[:, 1])
                                if sheet in col_data:
                                    self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' is present in TOC sheet.",
                                                                    pass_=True, log=True, screenshot=False)
                                else:
                                    self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' is not present in TOC sheet. "
                                                                            f"Available Data from TOC sheet: {col_data}",
                                                                    pass_=False, log=True, screenshot=False)
                                    raise Exception(f"'{sheet}' is present in TOC sheet which is not expected. "
                                                    f"Available Data from TOC sheet: {col_data}")
                                
                                qafile = pd.read_excel((os.getcwd()+stdyfile[index]))
                                excelfile = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name=sheet)

                                # Removing the 'Back To Toc' column to compare the exact data with uploaded file
                                excelfile = excelfile.iloc[:, 1:]
                                
                                if qafile.equals(excelfile):
                                    self.LogScreenshot.fLogScreenshot(message=f"File contents between QA File "
                                                                            f"'{Path(f'{(os.getcwd()+stdyfile[index])}').name}' and Complete Excel Report "
                                                                            f"'{Path(f'ActualOutputs//{excel_filename}').name}' "
                                                                            f"are matching",
                                                                    pass_=True, log=True, screenshot=False)
                                else:
                                    self.LogScreenshot.fLogScreenshot(message=f"File contents between QA File "
                                                                            f"'{Path({(os.getcwd()+stdyfile[index])}).name}' and Complete Excel Report "
                                                                            f"'{Path(f'ActualOutputs//{excel_filename}').name}' "
                                                                            f"are not matching",
                                                                    pass_=False, log=True, screenshot=False)
                                    raise Exception(f"File contents between QA File '{Path({(os.getcwd()+stdyfile[index])}).name}' "
                                                    f"and Complete Excel Report '{Path(f'ActualOutputs//{excel_filename}').name}' "
                                                    f"are not matching")
                            else:
                                raise Exception(f"'{sheet}' sheet is not present in complete excel report")
                    else:
                        if 'Quality Assessment' in excel_data.sheetnames:
                            self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' sheet is present in complete "
                                                                    f"excel report",
                                                            pass_=True, log=True, screenshot=False)

                            excel_sheet = excel_data['Quality Assessment']
                            if excel_sheet['A1'].value == 'Back To Toc':
                                self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is present",
                                                                pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(message=f"'Back To Toc' option is not present",
                                                                pass_=False, log=True, screenshot=False)
                                raise Exception(f"'Back To Toc' option is not present")
                            
                            toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                            col_data = list(toc_sheet.iloc[:, 1])
                            if 'Quality Assessment' in col_data:
                                self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' is present in TOC sheet.",
                                                                pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' is not present in TOC sheet. "
                                                                        f"Available Data from TOC sheet: {col_data}",
                                                                pass_=False, log=True, screenshot=False)
                                raise Exception(f"'Quality Assessment' is present in TOC sheet which is not expected. "
                                                f"Available Data from TOC sheet: {col_data}")
                            
                            qafile = pd.read_excel((os.getcwd()+stdyfile[0]))
                            excelfile = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="Quality Assessment")

                            # Removing the 'Back To Toc' column to compare the exact data with uploaded file
                            excelfile = excelfile.iloc[:, 1:]
                            
                            if qafile.equals(excelfile):
                                self.LogScreenshot.fLogScreenshot(message=f"File contents between QA File "
                                                                        f"'{Path(f'{(os.getcwd()+stdyfile[0])}').name}' and Complete Excel Report "
                                                                        f"'{Path(f'ActualOutputs//{excel_filename}').name}' "
                                                                        f"are matching",
                                                                pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(message=f"File contents between QA File "
                                                                        f"'{Path(f'{(os.getcwd()+stdyfile[0])}').name}' and Complete Excel Report "
                                                                        f"'{Path(f'ActualOutputs//{excel_filename}').name}' "
                                                                        f"are not matching",
                                                                pass_=False, log=True, screenshot=False)
                                raise Exception(f"File contents between QA File '{Path(f'{(os.getcwd()+stdyfile[0])}').name}' "
                                                f"and Complete Excel Report '{Path(f'ActualOutputs//{excel_filename}').name}' "
                                                f"are not matching")
                        else:
                            raise Exception("'Quality Assessment' sheet is not present in complete excel report")
        except Exception:
            raise Exception("Error in report comparision between QA data file and Complete Excel report")

    def del_data_after_qafile_comparison(self, locatorname, filepath, env):
        expected_delete_status_text = 'QA excel file successfully deleted'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)

        # Read slrtype data values
        slrtype = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        slrtype_ = sorted(list(set(tuple(sorted(sub)) for sub in slrtype)), key=lambda x: x[1])

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in slrtype_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["QA_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i]] for i in range(0, len(stdytype))]
                    
                    self.go_to_page("manage_qa_data_button", env)
                    for k in upload_data:
                        self.refreshpage()
                        time.sleep(3)
                        pop_value = self.base.selectbyvisibletext("select_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("select_stdy_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.LogScreenshot.fLogScreenshot(message=f'Selected Data For Deletion :',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("delete_file_button", env)
                        time.sleep(2)
                        self.click("delete_file_popup", env)
                        time.sleep(2)

                        actual_delete_status_text = self.get_text("get_status_text", env, UnivWaitFor=10)
                        # time.sleep(2)

                        if actual_delete_status_text == expected_delete_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'QA File Deletion is success for Population : '
                                                                    f'{pop_value} -> SLR Type : {selected_slr_val}.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f'Unable to find status message while deleting QA File '
                                                                    f'for Population : {pop_value} -> SLR Type : {selected_slr_val}.',
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while deleting QA File for Population : "
                                            f"{pop_value} -> SLR Type : {selected_slr_val}.")
                
                    # Go to live slr page
                    self.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(f"{i[0]}", f"{i[1]}", env)
                    self.slrreport.select_data(f"{j[0]}", f"{j[1]}", env)
                    self.slrreport.generate_download_report("excel_report", env)
                    excel_filename = self.slrreport.get_and_validate_filename(filepath)

                    excel_data = openpyxl.load_workbook(f'ActualOutputs//{excel_filename}')
                    if ['Clinical-Interventional', 'Clinical-RWE'] == stdytype:
                        expected_sheet_names = ['Quality Assessment-Intervtnl', 'Quality Assessment-RWE']
                        for sheet in expected_sheet_names:
                            if sheet not in excel_data.sheetnames:
                                self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' sheet is not present in complete "
                                                                        f"excel report as expected",
                                                                pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' sheet is present in complete "
                                                                        f"excel report which is not expected",
                                                                pass_=True, log=True, screenshot=False)
                                raise Exception(f"'{sheet}' sheet is present in complete excel report which is not "
                                                f"expected")
                            
                            toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                            col_data = list(toc_sheet.iloc[:, 1])
                            if sheet not in col_data:
                                self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' is not present in TOC sheet "
                                                                        f"as expected",
                                                                pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(message=f"'{sheet}' is present in TOC sheet which "
                                                                        f"is not expected. Available Data from "
                                                                        f"TOC sheet: {col_data}",
                                                                pass_=False, log=True, screenshot=False)
                                raise Exception(f"'{sheet}' is present in TOC sheet which is not expected. "
                                                f"Available Data from TOC sheet: {col_data}")
                    else:
                        if 'Quality Assessment' not in excel_data.sheetnames:
                            self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' sheet is not present in complete "
                                                                    f"excel report as expected",
                                                            pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' sheet is present in complete "
                                                                    f"excel report which is not expected",
                                                            pass_=True, log=True, screenshot=False)
                            raise Exception(f"'Quality Assessment' sheet is present in complete excel report which is not "
                                            f"expected")
                        
                        toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                        col_data = list(toc_sheet.iloc[:, 1])
                        if 'Quality Assessment' not in col_data:
                            self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' is not present in TOC sheet "
                                                                    f"as expected",
                                                            pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f"'Quality Assessment' is present in TOC sheet which "
                                                                    f"is not expected. Available Data from "
                                                                    f"TOC sheet: {col_data}",
                                                            pass_=False, log=True, screenshot=False)
                            raise Exception(f"'Quality Assessment' is present in TOC sheet which is not expected. "
                                            f"Available Data from TOC sheet: {col_data}")
        except Exception:
            raise Exception("Unable to delete the existing QA file")
