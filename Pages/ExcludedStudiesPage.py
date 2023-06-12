import enum
import os
from datetime import datetime
from pathlib import Path
import time
import openpyxl
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

from Pages.Base import Base
from Pages.ExtendedBasePage import ExtendedBase
from Pages.ManagePopulationsPage import ManagePopulationsPage
from Pages.OpenLiveSLRPage import LiveSLRPage
from Pages.SLRReportPage import SLRReport
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from utilities.logScreenshot import cLogScreenshot
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import Select


class ExcludedStudiesPage(Base):
    """Constructor of the ExcludedStudies Page class"""

    def __init__(self, driver, extra):
        # initializing the driver from base class
        super().__init__(driver, extra)
        self.extra = extra
        # Instantiate the Base class
        self.base = Base(self.driver, self.extra)
        # Creating object of ExtendedBase class
        self.exbase = ExtendedBase(self.driver, extra)        
        # Instantiate the logger class
        self.logger = LogGen.loggen()
        # Instantiate the logScreenshot class
        self.LogScreenshot = cLogScreenshot(self.driver, self.extra)
        # Creating object of liveslrpage class
        self.liveslrpage = LiveSLRPage(self.driver, extra)
        # Creating object of slrreport class
        self.slrreport = SLRReport(self.driver, extra)
        # Creating object of ManagePopulationsPage class
        self.mngpoppage = ManagePopulationsPage(self.driver, extra)
        # Instantiate webdriver wait class
        self.wait = WebDriverWait(driver, 10)

    # Reading slr study type and Excluded study files data for Excluded Publications Page -> upload feature validation
    def get_study_file_details(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        study_type = df.loc[df['Name'] == locatorname]['Study_Types'].dropna().to_list()
        qa_file = df.loc[df['Name'] == locatorname]['ExcludedStudies_Excel_Files'].dropna().to_list()
        filenames = df.loc[df['Name'] == locatorname]['ExcludedStudies_Excel_File_names'].dropna().to_list()
        result = [[study_type[i], os.getcwd() + qa_file[i], filenames[i]] for i in range(0, len(study_type))]
        return result

    # Reading slr study type and Excluded study files data for Excluded Publications Page -> override feature validation
    def get_study_file_details_override(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        study_type = df.loc[df['Name'] == locatorname]['Study_Types'].dropna().to_list()
        qa_file = df.loc[df['Name'] == locatorname]['Override_ExcludedStudies_Excel_Files'].dropna().to_list()
        filenames = df.loc[df['Name'] == locatorname]['Override_ExcludedStudies_Excel_File_names'].dropna().to_list()
        result = [(study_type[i], os.getcwd() + qa_file[i], filenames[i]) for i in range(0, len(study_type))]
        return result

    # Reading Population data for Excluded Publications Page
    def get_pop_data(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        pop = df.loc[df['Name'] == locatorname]['population_name'].dropna().to_list()
        pop_button = df.loc[df['Name'] == locatorname]['pop_radio_button'].dropna().to_list()
        result = [(pop[i], pop_button[i]) for i in range(0, len(pop))]
        return result

    # Reading slr study type and Excluded study files data for complete workflow validation
    def get_file_details(self, filepath):
        file = pd.read_excel(filepath)
        study_type = list(file['Study_Types'].dropna())
        study_file = list(os.getcwd() + file['ExcludedStudies_Excel_Files'].dropna())
        study_filename = list(file['ExcludedStudies_Excel_File_names'].dropna())
        result = [[study_type[i], study_file[i], study_filename[i]] for i in range(0, len(study_type))]
        return result

    # Check the presence of Manage Excluded Publications option in Admin page
    def presence_of_elements(self, locator, env):
        self.scroll(locator, env)
        self.wait.until(ec.presence_of_element_located((getattr(By, self.locatortype(locator, env)),
                                                        self.locatorpath(locator, env))))
        self.LogScreenshot.fLogScreenshot(message=f'Manage Excluded Publications option is present in Admin page.',
                                          pass_=True, log=True, screenshot=True)

    # Check Manage Excluded Publications page elements are accessible or not
    def access_excludedstudy_page_elements(self, locatorname, filepath, env):

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)
        # Read study types and file paths to upload
        stdy_data = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        stdy_data_ = sorted(list(set(tuple(sorted(sub)) for sub in stdy_data)), key=lambda x: x[1])

        df = pd.read_excel(filepath)

        for i in pop_data:
            for j in stdy_data_:
                # Get StudyType and Files path to upload Managae QA Data
                data1 = df[df["Name"] == locatorname]
                data1_val = data1[data1["slrtype"] == j[0]]
                stdytype = data1_val["Study_Types"]
                stdytype = [item for item in stdytype if str(item) != 'nan']
                stdyfile = data1_val["ExcludedStudies_Excel_Files"]
                stdyfile = [item for item in stdyfile if str(item) != 'nan']
                stdyfile_name = data1_val["ExcludedStudies_Excel_File_names"]
                stdyfile_name = [item for item in stdyfile_name if str(item) != 'nan']

                upload_data = [[stdytype[i], os.getcwd()+stdyfile[i], stdyfile_name[i]] for i in range(0, len(stdytype))]

                self.go_to_page("excluded_studies_link", env)
                for k in upload_data:
                    self.refreshpage()
                    time.sleep(3)
                    self.click("ex_stdy_pop_dropdown", env)
                    self.LogScreenshot.fLogScreenshot(message=f"Population dropdown is accessible. Listed elements are:",
                                                    pass_=True, log=True, screenshot=True)
                    selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                    time.sleep(1)

                    self.click("ex_stdy_stdytype_dropdown", env)
                    self.LogScreenshot.fLogScreenshot(message=f"SLR Type dropdown is accessible. Listed elements are:",
                                                    pass_=True, log=True, screenshot=True)
                    selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                    time.sleep(1)

                    '''Commenting this section due to LIVEHTA-2980 implementation'''
                    # selected_update_val = self.base.selectbyindex("ex_stdy_update_dropdown", 1, env)
                    # expected_table_values.append(selected_update_val)
                    # time.sleep(1)

                    '''Below JavaScript command is to manipulate the input type='file' tag to make it visible for selenium to upload the file'''
                    cmd = "document.getElementsByTagName('style')[7].textContent='.file-input[_ngcontent-org-c124]" \
                        "{display:visible}'"
                    self.jsclick_hide(cmd)
                    time.sleep(1)

                    self.input_text("ex_stdy_file_upload", k[1], env)
                    time.sleep(2)

                    if self.clickable("ex_stdy_upload_button", env):
                        self.LogScreenshot.fLogScreenshot(message=f"Upload button is clickable after selecting the file",
                                                        pass_=True, log=True, screenshot=True)
                    else:
                        self.LogScreenshot.fLogScreenshot(message=f"Upload button is not clickable after selecting the "
                                                                f"file",
                                                        pass_=False, log=True, screenshot=True)
                        raise Exception("Upload button is not clickable after selecting the file")

    def add_multiple_excluded_study_data(self, locatorname, filepath, env):
        expected_upload_status_text = 'File(s) uploaded successfully'
        # Read the username
        username = self.get_text("get_user_name", env, UnivWaitFor=10)
        firstname = username.split()[0]

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)
        # Read study types and file paths to upload
        stdy_data = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        stdy_data_ = sorted(list(set(tuple(sorted(sub)) for sub in stdy_data)), key=lambda x: x[1])

        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["ExcludedStudies_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']
                    stdyfile_name = data1_val["ExcludedStudies_Excel_File_names"]
                    stdyfile_name = [item for item in stdyfile_name if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i], stdyfile_name[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("excluded_studies_link", env)
                    for k in upload_data:
                        expected_table_values = []
                        self.refreshpage()
                        time.sleep(3)
                        selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        expected_table_values.append(selected_pop_val)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        expected_table_values.append(j[0])
                        time.sleep(1)

                        '''Commenting this section due to LIVEHTA-2980 implementation'''
                        # selected_update_val = self.base.selectbyindex("ex_stdy_update_dropdown", 1, env)
                        # expected_table_values.append(selected_update_val)
                        # time.sleep(1)

                        '''Below JavaScript command is to manipulate the input type='file' tag to make it visible for selenium to upload the file'''
                        cmd = "document.getElementsByTagName('style')[7].textContent='.file-input[_ngcontent-org-c124]" \
                            "{display:visible}'"
                        self.jsclick_hide(cmd)
                        time.sleep(1)

                        self.input_text("ex_stdy_file_upload", k[1], env)
                        expected_table_values.append(k[2])
                        time.sleep(2)

                        self.click("ex_stdy_upload_button", env)
                        time.sleep(4)
                        actual_upload_status_text = self.get_status_text("ex_stdy_status_text", env, UnivWaitFor=10)

                        if actual_upload_status_text == expected_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'Excluded Publications File upload is success for '
                                                                    f'Population : {i[0]} -> SLR Type : {k[0]}.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f'Unable to find status message while uploading '
                                                                    f'Excluded Publications File for Population : '
                                                                    f'{i[0]} -> SLR Type: {k[0]}.',
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception("Unable to find status message during Excluded Publications file uploading")

                        # Add the firstname to expected values list
                        expected_table_values.append(firstname)

                        selected_pop_val1 = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val1 = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        time.sleep(1)

                        actual_table_values = self.get_texts('ex_stdy_table_data_row_1', env)

                        for n in expected_table_values:
                            if n in actual_table_values:
                                self.LogScreenshot.fLogScreenshot(message=f'Correct data is present in table.',
                                                                pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(message=f'Mismatch is found in table data. '
                                                                        f'Expected values are : {expected_table_values} '
                                                                        f'and Actual values are : {actual_table_values}',
                                                                pass_=False, log=True, screenshot=True)
                                raise Exception(f'Mismatch is found in table data. Expected values are : '
                                                f'{expected_table_values} and Actual values are : {actual_table_values}')
        except Exception:
            raise Exception("Unable to upload the Excluded Publications data")

    def update_multiple_excluded_study_data(self, locatorname, filepath, env):
        expected_upload_status_text = 'File(s) uploaded successfully'
        # Read the username
        username = self.get_text("get_user_name", env, UnivWaitFor=10)
        firstname = username.split()[0]

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)
        # Read study types and file paths to upload
        stdy_data = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        stdy_data_ = sorted(list(set(tuple(sorted(sub)) for sub in stdy_data)), key=lambda x: x[1])

        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["Override_ExcludedStudies_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']
                    stdyfile_name = data1_val["Override_ExcludedStudies_Excel_File_names"]
                    stdyfile_name = [item for item in stdyfile_name if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i], stdyfile_name[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("excluded_studies_link", env)
                    for k in upload_data:
                        expected_table_values = []
                        self.refreshpage()
                        time.sleep(3)
                        selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        expected_table_values.append(selected_pop_val)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        expected_table_values.append(j[0])
                        time.sleep(1)

                        '''Commenting this section due to LIVEHTA-2980 implementation'''
                        # selected_update_val = self.base.selectbyindex("ex_stdy_update_dropdown", 1, env)
                        # expected_table_values.append(selected_update_val)
                        # time.sleep(1)

                        '''Below JavaScript command is to manipulate the input type='file' tag to make it visible for selenium to upload the file'''
                        cmd = "document.getElementsByTagName('style')[7].textContent='.file-input[_ngcontent-org-c124]" \
                            "{display:visible}'"
                        self.jsclick_hide(cmd)
                        time.sleep(1)

                        self.input_text("ex_stdy_file_upload", k[1], env)
                        expected_table_values.append(k[2])
                        time.sleep(2)

                        self.click("ex_stdy_upload_button", env)
                        time.sleep(3)
                        self.LogScreenshot.fLogScreenshot(
                            message=f"Confirmation popup while updating the existing Excluded Publications File.",
                            pass_=True, log=True, screenshot=True)
                        self.jsclick("ex_stdy_popup_ok", env, message="Expected : popup reminder. Actual : popup is "
                                                                    "not shown")
                        time.sleep(3)
                        # actual_upload_status_text = self.get_text("ex_stdy_status_text", env, UnivWaitFor=10)
                        actual_upload_status_text = self.get_status_text("ex_stdy_status_text", env, UnivWaitFor=10)
                        # time.sleep(1)

                        if actual_upload_status_text == expected_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(
                                message=f'For Population : {i[0]} -> SLR Type : {k[0]}, updating the existing Excluded '
                                        f'Studies File is success.',
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f'For Population : {i[0]} -> SLR Type : {k[0]}, Unable to find status message '
                                        f'while updating the existing Excluded Publications File.',
                                pass_=False, log=True, screenshot=True)
                            raise Exception(
                                "Unable to find status message while Updating the existing Excluded Publications "
                                "file")

                        # Add the firstname to expected values list
                        expected_table_values.append(firstname)

                        selected_pop_val1 = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val1 = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        time.sleep(1)

                        actual_table_values = self.get_texts('ex_stdy_table_data_row_1', env)

                        for n in expected_table_values:
                            if n in actual_table_values:
                                self.LogScreenshot.fLogScreenshot(message=f'Updated data is present in table.',
                                                                pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f'Mismatch is found in table data. Expected values are : '
                                            f'{expected_table_values} and Actual values are : {actual_table_values}',
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(
                                    f'Mismatch is found in table data. Expected values are : {expected_table_values} '
                                    f'and Actual values are : {actual_table_values}')
        except Exception:
            raise Exception("Unable to update the existing Excluded Publications data")

    def del_multiple_excluded_study_data(self, locatorname, filepath, env):
        expected_delete_status_text = 'Excluded publications file deleted successfully'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)
        # Read study types and file paths to upload
        stdy_data = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        stdy_data_ = sorted(list(set(tuple(sorted(sub)) for sub in stdy_data)), key=lambda x: x[1])

        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["ExcludedStudies_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']
                    stdyfile_name = data1_val["ExcludedStudies_Excel_File_names"]
                    stdyfile_name = [item for item in stdyfile_name if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i], stdyfile_name[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("excluded_studies_link", env)
                    for k in upload_data:
                        expected_table_values = []
                        self.refreshpage()
                        time.sleep(3)
                        selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        expected_table_values.append(selected_pop_val)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        expected_table_values.append(selected_slr_val)
                        time.sleep(1)

                        '''Commenting this section due to LIVEHTA-2980 implementation'''
                        # selected_update_val = self.base.selectbyindex("ex_stdy_update_dropdown", 1, env)
                        # expected_table_values.append(selected_update_val)
                        # time.sleep(1)

                        self.LogScreenshot.fLogScreenshot(message=f'Data selected for deletion is : ',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("ex_stdy_delete", env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(
                            message=f"Confirmation popup while deleting the Excluded Publications File.",
                            pass_=True, log=True, screenshot=True)
                        self.click("ex_stdy_popup_ok", env)
                        time.sleep(2)

                        # actual_delete_status_text = self.get_text("ex_stdy_status_text", env, UnivWaitFor=10)
                        actual_delete_status_text = self.get_status_text("ex_stdy_status_text", env, UnivWaitFor=10)
                        # time.sleep(2)

                        if actual_delete_status_text == expected_delete_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'Excluded Publications File Deletion is success.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f'Unable to find status message while deleting Excluded Publications File.',
                                pass_=False, log=True, screenshot=True)
                            raise Exception("Error in Excluded Publications File Deletion")
        except Exception:
            raise Exception("Unable to delete the existing Excluded Publications File")

    def compare_excludedstudy_file_with_report(self, filepath, locatorname, env, prj_name):
        expected_upload_status_text = 'File(s) uploaded successfully'
        # Read the username
        username = self.get_text("get_user_name", env, UnivWaitFor=10)
        firstname = username.split()[0]

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)
        # Read study types and file paths to upload
        stdy_data = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        stdy_data_ = sorted(list(set(tuple(sorted(sub)) for sub in stdy_data)), key=lambda x: x[1])

        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data_:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["ExcludedStudies_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']
                    stdyfile_name = data1_val["ExcludedStudies_Excel_File_names"]
                    stdyfile_name = [item for item in stdyfile_name if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i], stdyfile_name[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("excluded_studies_link", env)
                    for k in upload_data:
                        expected_table_values = []
                        self.refreshpage()
                        time.sleep(3)
                        selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        expected_table_values.append(selected_pop_val)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        expected_table_values.append(j[0])
                        time.sleep(1)

                        '''Commenting this section due to LIVEHTA-2980 implementation'''
                        # selected_update_val = self.base.selectbyindex("ex_stdy_update_dropdown", 1, env)
                        # expected_table_values.append(selected_update_val)
                        # time.sleep(1)

                        '''Below JavaScript command is to manipulate the input type='file' tag to make it visible for selenium to upload the file'''
                        cmd = "document.getElementsByTagName('style')[7].textContent='.file-input[_ngcontent-org-c124]" \
                            "{display:visible}'"
                        self.jsclick_hide(cmd)
                        time.sleep(1)

                        self.input_text("ex_stdy_file_upload", k[1], env)
                        expected_table_values.append(k[2])
                        time.sleep(2)

                        self.click("ex_stdy_upload_button", env)
                        time.sleep(2)
                        actual_upload_status_text = self.get_status_text("ex_stdy_status_text", env, UnivWaitFor=10)
                        # time.sleep(1)

                        if actual_upload_status_text == expected_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(
                                message=f'Excluded Publications File upload is success for Population : {i[0]} -> '
                                        f'SLR Type : {k[0]}.',
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f'Unable to find status message while uploading Excluded Publications File for '
                                        f'Population : {i[0]} -> SLR Type: {k[0]}.',
                                pass_=False, log=True, screenshot=True)
                            raise Exception("Unable to find status message during Excluded Publications file uploading")

                        # Add the firstname to expected values list
                        expected_table_values.append(firstname)

                        # Read table data for specific population and slr study type
                        selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        time.sleep(2)

                        actual_table_values = self.get_texts('ex_stdy_table_data_row_1', env)

                        for n in expected_table_values:
                            if n in actual_table_values:
                                self.LogScreenshot.fLogScreenshot(message=f'Correct data is present in table.',
                                                                pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f'Mismatch is found in table data. Expected values are : '
                                            f'{expected_table_values} and Actual values are : {actual_table_values}',
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(
                                    f'Mismatch is found in table data. Expected values are : {expected_table_values} and '
                                    f'Actual values are : {actual_table_values}')

                    # Go to live slr page
                    self.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(f"{i[0]}", f"{i[1]}", env)
                    self.slrreport.select_data(f"{j[0]}", f"{j[1]}", env)
                    self.slrreport.generate_download_report("excel_report", env)
                    excel_filename = self.slrreport.get_and_validate_filename(filepath)
                    
                    # if prj_name == "Oncology":
                    #     date_val = expected_table_values[2].translate({ord('/'): None})[-4:] + \
                    #                expected_table_values[2].translate({ord('/'): None})[:4]
                    #     excluded_sheet_name = f"Excluded studies {date_val}"
                    # else:
                    #     date_val = datetime.strptime(expected_table_values[2], '%m/%d/%Y').date()
                    #     excluded_sheet_name = f"Excluded Pubs-{date_val}"

                    excel_data = openpyxl.load_workbook(f'ActualOutputs//{excel_filename}')
                    if ['Clinical -Interventional', 'Clinical -RWE'] == stdytype:
                        expected_sheet_names = ['Excluded Pubs-Intervtnl', 'Excluded Pubs-RWE']
                        for index, sheet in enumerate(expected_sheet_names):
                            if sheet in excel_data.sheetnames:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'{sheet}' sheet is present in complete excel report",
                                    pass_=True, log=True, screenshot=False)

                                excel_sheet = excel_data[f'{sheet}']
                                if excel_sheet['A1'].value == 'Back To Toc':
                                    self.LogScreenshot.fLogScreenshot(
                                        message=f"'Back To Toc' option is present in '{sheet}' "
                                                f"sheet",
                                        pass_=True, log=True, screenshot=False)
                                else:
                                    self.LogScreenshot.fLogScreenshot(
                                        message=f"'Back To Toc' option is not present in '{sheet}' "
                                                f"sheet",
                                        pass_=False, log=True, screenshot=False)
                                    raise Exception(
                                        f"'Back To Toc' option is not present in '{sheet}' sheet")

                                toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                                col_data = list(toc_sheet.iloc[:, 1])
                                if f'{sheet}' in col_data:
                                    self.LogScreenshot.fLogScreenshot(
                                        message=f"'{sheet}' is present in TOC sheet.",
                                        pass_=True, log=True, screenshot=False)
                                else:
                                    self.LogScreenshot.fLogScreenshot(
                                        message=f"'{sheet}' is not present in TOC sheet. Available Data from "
                                                f"TOC sheet: '{col_data}'", pass_=False, log=True, screenshot=False)
                                    raise Exception(f"'{sheet}' is not present in TOC sheet.")

                                studyfile = pd.read_excel((os.getcwd()+stdyfile[index]))
                                excelfile = pd.read_excel(f'ActualOutputs//{excel_filename}',
                                                        sheet_name=f'{sheet}')

                                # Removing the 'Back To Toc' column to compare the exact data with uploaded file
                                excelfile = excelfile.iloc[:, 1:]

                                if studyfile.equals(excelfile):
                                    self.LogScreenshot.fLogScreenshot(
                                        message=f"File contents between QA File '{Path(f'{(os.getcwd()+stdyfile[index])}').name}' and Complete Excel "
                                                f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are matching",
                                        pass_=True, log=True, screenshot=False)
                                else:
                                    raise Exception(
                                        f"File contents between Study File '{Path(f'{(os.getcwd()+stdyfile[index])}').name}' and Complete Excel Report "
                                        f"'{Path(f'ActualOutputs//{excel_filename}').name}' are not matching")
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'{sheet}' sheet is not present in complete excel report",
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(f"'{sheet}' sheet is not present in complete excel report")
                    else:
                        if 'Excluded Pubs-Original SLR' in excel_data.sheetnames:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"'Excluded Pubs-Original SLR' sheet is present in complete excel report",
                                pass_=True, log=True, screenshot=False)

                            excel_sheet = excel_data[f'Excluded Pubs-Original SLR']
                            if excel_sheet['A1'].value == 'Back To Toc':
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'Back To Toc' option is present in 'Excluded Pubs-Original SLR' "
                                            f"sheet",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'Back To Toc' option is not present in 'Excluded Pubs-Original SLR' "
                                            f"sheet",
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(
                                    f"'Back To Toc' option is not present in 'Excluded Pubs-Original SLR' sheet")

                            toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                            col_data = list(toc_sheet.iloc[:, 1])
                            if f'Excluded Pubs-Original SLR' in col_data:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'Excluded Pubs-Original SLR' is present in TOC sheet.",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'Excluded Pubs-Original SLR' is not present in TOC sheet. Available Data from "
                                            f"TOC sheet: '{col_data}'", pass_=False, log=True, screenshot=False)
                                raise Exception(f"'Excluded Pubs-Original SLR' is not present in TOC sheet.")

                            studyfile = pd.read_excel((os.getcwd()+stdyfile[0]))
                            excelfile = pd.read_excel(f'ActualOutputs//{excel_filename}',
                                                    sheet_name=f'Excluded Pubs-Original SLR')

                            # Removing the 'Back To Toc' column to compare the exact data with uploaded file
                            excelfile = excelfile.iloc[:, 1:]

                            if studyfile.equals(excelfile):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"File contents between QA File '{Path(f'{(os.getcwd()+stdyfile[index])}').name}' and Complete Excel "
                                            f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are matching",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                raise Exception(
                                    f"File contents between Study File '{Path(f'{(os.getcwd()+stdyfile[index])}').name}' and Complete Excel Report "
                                    f"'{Path(f'ActualOutputs//{excel_filename}').name}' are not matching")
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"'Excluded Pubs-Original SLR' sheet is not present in complete excel report",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(f"'Excluded Pubs-Original SLR' sheet is not present in complete excel report")
        except Exception:
            raise Exception("Error in report comparision between Excluded study file and Complete Excel report")

    def del_after_studyfile_comparison(self, filepath, locatorname, env, prj_name):
        expected_delete_status_text = 'Excluded publications file deleted successfully'

        # Read population details from data sheet
        pop_data = self.exbase.get_population_data(filepath, 'Sheet1', locatorname)
        # Read study types and file paths to upload
        stdy_data = self.exbase.get_slrtype_data(filepath, 'Sheet1', locatorname)
        # Removing duplicates to get the proper length of SLR Type data
        stdy_data_ = sorted(list(set(tuple(sorted(sub)) for sub in stdy_data)), key=lambda x: x[1])

        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data_:                    
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    stdyfile = data1_val["ExcludedStudies_Excel_Files"]
                    stdyfile = [item for item in stdyfile if str(item) != 'nan']
                    stdyfile_name = data1_val["ExcludedStudies_Excel_File_names"]
                    stdyfile_name = [item for item in stdyfile_name if str(item) != 'nan']

                    upload_data = [[stdytype[i], os.getcwd()+stdyfile[i], stdyfile_name[i]] for i in range(0, len(stdytype))]

                    self.go_to_page("excluded_studies_link", env)
                    for k in upload_data:
                        expected_table_values = []
                        self.refreshpage()
                        time.sleep(3)
                        selected_pop_val = self.base.selectbyvisibletext("ex_stdy_pop_dropdown", i[0], env)
                        expected_table_values.append(selected_pop_val)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("ex_stdy_stdytype_dropdown", k[0], env)
                        expected_table_values.append(selected_slr_val)
                        time.sleep(1)

                        '''Commenting this section due to LIVEHTA-2980 implementation'''
                        # selected_update_val = self.base.selectbyindex("ex_stdy_update_dropdown", 1, env)
                        # expected_table_values.append(selected_update_val)
                        # time.sleep(1)

                        self.LogScreenshot.fLogScreenshot(message=f'Data selected for deletion is : ',
                                                        pass_=True, log=True, screenshot=True)

                        self.click("ex_stdy_delete", env)
                        time.sleep(2)
                        self.LogScreenshot.fLogScreenshot(
                            message=f"Confirmation popup while deleting the Excluded Publications File.",
                            pass_=True, log=True, screenshot=True)
                        self.click("ex_stdy_popup_ok", env)
                        time.sleep(2)

                        actual_delete_status_text = self.get_status_text("ex_stdy_status_text", env, UnivWaitFor=10)
                        # time.sleep(2)                   

                        if actual_delete_status_text == expected_delete_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f'Excluded Publications File Deletion is success.',
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f'Error while deleting Excluded Publications File. Error Message is '
                                        f'{actual_delete_status_text}',
                                pass_=False, log=True, screenshot=True)
                            raise Exception("Error in Excluded Publications File Deletion")

                    # Go to live slr page
                    self.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(f"{i[0]}", f"{i[1]}", env)
                    self.slrreport.select_data(f"{j[0]}", f"{j[1]}", env)
                    self.slrreport.generate_download_report("excel_report", env)
                    excel_filename = self.slrreport.get_and_validate_filename(filepath)

                    # if prj_name == "Oncology":
                    #     date_val = expected_table_values[2].translate({ord('/'): None})[-4:] + \
                    #                expected_table_values[2].translate({ord('/'): None})[:4]
                    #     excluded_sheet_name = f"Excluded studies {date_val}"
                    # else:
                    #     date_val = datetime.strptime(expected_table_values[2], '%m/%d/%Y').date()
                    #     excluded_sheet_name = f"Excluded Pubs-{date_val}"

                    excel_data = openpyxl.load_workbook(f'ActualOutputs//{excel_filename}')
                    if ['Clinical -Interventional', 'Clinical -RWE'] == stdytype:
                        expected_sheet_names = ['Excluded Pubs-Intervtnl', 'Excluded Pubs-RWE']
                        for sheet in expected_sheet_names:
                            if sheet not in excel_data.sheetnames:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'{sheet}' sheet is not present in complete excel report as expected",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"'{sheet}' sheet is present in complete excel report which is not "
                                            f"expected", pass_=False, log=True, screenshot=False)
                                raise Exception(
                                    f"'{sheet}' sheet is present in complete excel report which is not expected")

                            toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                            col_data = list(toc_sheet.iloc[:, 1])
                            if sheet not in col_data:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f'{sheet} is not present in TOC sheet as expected.',
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f'{sheet} is present in TOC sheet which is not expected. '
                                            f'Available Data from TOC sheet: {col_data}',
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(f"'{sheet}' is present in TOC sheet.")
                    else:
                        if 'Excluded Pubs-Original SLR' not in excel_data.sheetnames:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"'Excluded Pubs-Original SLR' sheet is not present in complete excel report as expected",
                                pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"'Excluded Pubs-Original SLR' sheet is present in complete excel report which is not "
                                        f"expected", pass_=False, log=True, screenshot=False)
                            raise Exception(
                                f"'Excluded Pubs-Original SLR' sheet is present in complete excel report which is not expected")

                        toc_sheet = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name="TOC", skiprows=3)
                        col_data = list(toc_sheet.iloc[:, 1])
                        if 'Excluded Pubs-Original SLR' not in col_data:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"'Excluded Pubs-Original SLR' is not present in TOC sheet as expected.",
                                pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"'Excluded Pubs-Original SLR' is present in TOC sheet which is not expected. "
                                        f"Available Data from TOC sheet: {col_data}",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(f"'Excluded Pubs-Original SLR' is present in TOC sheet.")
        except Exception:
            raise Exception("Unable to delete the existing Excluded Publications File")
