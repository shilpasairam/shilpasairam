from datetime import date, datetime, timedelta
import os
import time
import pandas as pd
import numpy as np
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait

from pathlib import Path
from Pages.Base import Base
from Pages.ExtendedBasePage import ExtendedBase
from Pages.SLRReportPage import SLRReport
from utilities.customLogger import LogGen
from utilities.logScreenshot import cLogScreenshot
from selenium.webdriver.support.ui import Select

from utilities.readProperties import ReadConfig


class ProtocolPage(Base):

    """Constructor of the Protocol Page class"""
    def __init__(self, driver, extra):
        # initializing the driver from base class
        super().__init__(driver, extra)
        self.extra = extra
        # Instantiate the Base class
        self.base = Base(self.driver, self.extra)
        # Creating object of ExtendedBase class
        self.exbase = ExtendedBase(self.driver, extra)
        # Creating object of slrreport class
        self.slrreport = SLRReport(self.driver, extra)
        # Instantiate the logger class
        self.logger = LogGen.loggen()
        # Instantiate the logScreenshot class
        self.LogScreenshot = cLogScreenshot(self.driver, self.extra)
        # Instantiate webdriver wait class
        self.wait = WebDriverWait(driver, 10)

    # Reading Population data for PRISMAs Page
    def get_prisma_pop_data(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        pop = df.loc[df['Name'] == locatorname]['Prisma_Population'].dropna().to_list()
        return pop

    def get_prisma_excelfile_details(self, filepath, locatorname, column_name):
        df = pd.read_excel(filepath)
        sheet_name = os.getcwd()+(df.loc[df['Name'] == locatorname][column_name].to_list()[0])
        return sheet_name

    def get_prisma_data_details(self, filepath, locatorname):
        df = pd.read_excel(filepath)
        sheet_name = df.loc[df['Name'] == locatorname]['Study_Types'].to_list()
        sheet1 = df.loc[df['Name'] == locatorname]['OriginalStudiesNumbers'].to_list()
        sheet2 = df.loc[df['Name'] == locatorname]['RecordsNumber'].to_list()
        sheet3 = df.loc[df['Name'] == locatorname]['fullTextReviewRecordsNumber'].to_list()
        sheet4 = df.loc[df['Name'] == locatorname]['totalRecordsNumber'].to_list()
        sheet5 = df.loc[df['Name'] == locatorname]['Prisma_Image'].to_list()
        prisma_data = [(sheet_name[i], int(sheet1[i]), int(sheet2[i]), int(sheet3[i]), int(sheet4[i]),
                        (os.getcwd()+sheet5[i])) for i in range(0, len(sheet_name))]
        return prisma_data

    def add_prisma_excel_file(self, locatorname, filepath, env):
        expected_excel_upload_status_text = "PRISMA successfully uploaded"

        # Read population details from data sheet
        pop_name = self.get_prisma_pop_data(filepath, locatorname)

        try:
            selected_pop_val = self.base.selectbyvisibletext("prisma_pop_dropdown", pop_name[0], env)

            # Read the PRISMA file path required to upload
            prisma_excel = self.get_prisma_excelfile_details(filepath, locatorname, 'Prisma_Excel_File')

            self.input_text("prisma_excel_file", prisma_excel, env, UnivWaitFor=10)
            self.click("prisma_excel_upload_btn", env)
            time.sleep(2)
            
            actual_excel_upload_status_text = self.get_status_text("prisma_excel_status_text", env)

            if actual_excel_upload_status_text == expected_excel_upload_status_text:
                self.LogScreenshot.fLogScreenshot(message=f"PRISMA Excel File Upload is success.",
                                                  pass_=True, log=True, screenshot=True)
            else:
                self.LogScreenshot.fLogScreenshot(message=f"Unable to find status message while uploading "
                                                          f"PRISMA Excel File. Actual status message is "
                                                          f"{actual_excel_upload_status_text} and Expected status "
                                                          f"message is {expected_excel_upload_status_text}",
                                                  pass_=False, log=True, screenshot=True)
                raise Exception("Unable to find status message while uploading PRISMA Excel File.")
        except Exception:
            raise Exception("Unable to upload PRISMA Excel file")

    def del_prisma_excel_file(self, locatorname, excel_del_locator, excel_del_popup, filepath, env):
        expected_excel_del_status_text = "PRISMA excel file successfully deleted"

        # Read population details from data sheet
        pop_name = self.get_prisma_pop_data(filepath, locatorname)

        self.refreshpage()
        time.sleep(3)
        try:
            selected_pop_val = self.base.selectbyvisibletext("prisma_pop_dropdown", pop_name[0], env)

            self.del_prisma_image(locatorname, filepath, "prisma_image_delete_btn", "prisma_image_delete_popup", env)
            time.sleep(2)

            self.click(excel_del_locator, env)
            time.sleep(2)
            self.click(excel_del_popup, env)
            time.sleep(2)

            actual_excel_del_status_text = self.get_status_text("prisma_excel_status_text", env)

            if actual_excel_del_status_text == expected_excel_del_status_text:
                self.LogScreenshot.fLogScreenshot(message=f"PRISMA Excel File Delete is success.",
                                                  pass_=True, log=True, screenshot=True)
            else:
                self.LogScreenshot.fLogScreenshot(message=f"Unable to find status message while deleting "
                                                          f"PRISMA Excel File. Actual status message is "
                                                          f"{actual_excel_del_status_text} and Expected status "
                                                          f"message is {expected_excel_del_status_text}",
                                                  pass_=False, log=True, screenshot=True)
                raise Exception("Unable to find status message while deleting PRISMA Excel File.")
            time.sleep(5)
        except Exception:
            raise Exception("Unable to delete PRISMA Excel file")

    def upload_prisma_excel(self, locatorname, filepath, pop_data, stdy_data, env, project, fname):
        expected_excel_upload_status_text = "PRISMA successfully updated"

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)
        # Read the Data required to upload for study types
        # stdy_data = self.exbase.get_double_col_data(filepath, locatorname, 'Sheet1', 'Study_Types', 'Prisma_Image')
        prisma_numbers = self.exbase.get_double_col_data(filepath, locatorname, 'Sheet1', 'stdy_type_locators',
                                                         'stdy_type_values')

        try:
            for i in pop_data:
                for j in stdy_data:
                    # Get StudyType and Files path to upload PRISMA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    excel = data1_val["Prisma_Excel_File"]
                    excel = [item for item in excel if str(item) != 'nan']
                    template = data1_val["Template_name_prisma"]
                    template = [item for item in template if str(item) != 'nan']

                    upload_data = [[stdytype[i], excel[i], template[i]] for i in range(0, len(stdytype))]

                    for k in upload_data:
                        self.refreshpage()
                        selected_pop_val = self.base.selectbyvisibletext("prisma_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("prisma_study_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.slrreport.generate_download_report("prisma_template_download_btn", env)
                        # Renaming the filename because there is an issue in downloading filenames with same name
                        # multiple times in headless mode. So as an alternative renaming the file after downloading
                        # in each iteration
                        self.file_rename(self.slrreport.get_latest_filename(UnivWaitFor=180),
                                         f"{k[0]}_PRISMA-File-Template_{project}_{fname}.xlsx")
                        expected_template_name = f"{k[2]}_{fname}.xlsx"
                        downloaded_template_name = self.slrreport.get_latest_filename(UnivWaitFor=180)
                        if downloaded_template_name == expected_template_name:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Correct Template is downloaded. Template name is {downloaded_template_name}",
                                pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Mismatch in search strategy template name. Expected Template name is "
                                        f"{expected_template_name} and Actual Template name is "
                                        f"{downloaded_template_name}",
                                pass_=False, log=True, screenshot=False)

                        for num in prisma_numbers:
                            self.input_text(num[0], int(num[1]), env, UnivWaitFor=5)

                        self.input_text("prisma_image", os.getcwd()+"\\"+k[1], env, UnivWaitFor=5)
                        self.click("prisma_image_save_btn", env)
                        time.sleep(3)

                        actual_excel_upload_status_text = self.get_status_text("prisma_image_status_text", env)

                        if actual_excel_upload_status_text == expected_excel_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f"PRISMA Excel File Upload is success for Population '{i[0]}' -> SLR Type '{k[0]}'.",
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f"Unable to find status message while uploading "
                                                                    f"PRISMA Excel File for Population '{i[0]}' -> SLR Type '{k[0]}'. Actual status message is "
                                                                    f"{actual_excel_upload_status_text} and Expected "
                                                                    f"status message is {expected_excel_upload_status_text}",
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception("Unable to find status message while uploading PRISMA Excel File.")
        except Exception:
            raise Exception("Unable to upload PRISMA Excel")

    def del_prisma_excel(self, locatorname, filepath, pop_data, stdy_data, env):
        expected_excel_del_status_text = "PRISMA successfully deleted"

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data:
                    # Get StudyType and Files path to upload PRISMA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']

                    upload_data = [[stdytype[i]] for i in range(0, len(stdytype))]

                    for k in upload_data:
                        self.refreshpage()
                        selected_pop_val = self.base.selectbyvisibletext("prisma_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("prisma_study_type_dropdown", k[0], env)
                        time.sleep(1)

                        self.click("prisma_image_delete_btn", env)
                        time.sleep(2)
                        self.click("prisma_image_delete_popup", env)
                        time.sleep(2)

                        actual_excel_del_status_text = self.get_status_text("prisma_image_status_text", env)

                        if actual_excel_del_status_text == expected_excel_del_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f"PRISMA Excel File Delete is success for Population '{i[0]}' -> SLR Type '{k[0]}'.",
                                                            pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f"Unable to find status message while deleting "
                                                                    f"PRISMA Excel File for Population '{i[0]}' -> SLR Type '{k[0]}'. Actual status message is "
                                                                    f"{actual_excel_del_status_text} and Expected status "
                                                                    f"message is {expected_excel_del_status_text}",
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception("Unable to find status message while deleting PRISMA Excel File.")
                        time.sleep(5)
        except Exception:
            raise Exception("Unable to delete PRISMA Excel")

    def override_prisma_details(self, locatorname, filepath, index, pop_data, stdy_data, env):
        expected_excel_upload_status_text = "There is an existing PRISMA file for this population. " \
                                            "Please delete the existing file before uploading a new one"

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)
        prisma_numbers = self.exbase.get_double_col_data(filepath, locatorname, 'Sheet1', 'stdy_type_locators',
                                                         'stdy_type_values')

        try:
            for i in pop_data:
                for j in stdy_data:
                    # Get StudyType and Files path to upload PRISMA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    excel = data1_val["Prisma_Excel_File"]
                    excel = [item for item in excel if str(item) != 'nan']
                    template = data1_val["Template_name_prisma"]
                    template = [item for item in template if str(item) != 'nan']

                    upload_data = [[stdytype[i], excel[i], template[i]] for i in range(0, len(stdytype))]

                    for k in upload_data:
                        self.refreshpage()
                        selected_pop_val = self.base.selectbyvisibletext("prisma_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("prisma_study_type_dropdown", k[0], env)
                        time.sleep(1)

                        for num in prisma_numbers:
                            self.input_text(num[0], int(num[1])+index, env, UnivWaitFor=5)

                        self.input_text("prisma_image", os.getcwd()+"\\"+k[1], env, UnivWaitFor=5)
                        self.click("prisma_image_save_btn", env)
                        time.sleep(2)

                        actual_excel_upload_status_text = self.get_status_text("prisma_excel_status_text", env)

                        if actual_excel_upload_status_text == expected_excel_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(message=f"There is an existing PRISMA Excel file for Population : "
                                                                    f"'{i[0]}' -> SLR Type '{k[0]}'", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(message=f"Unable to find status message while overriding "
                                                                    f"PRISMA Excel File for Population '{i[0]}' -> SLR Type '{k[0]}'. Actual status message is "
                                                                    f"{actual_excel_upload_status_text} and Expected status "
                                                                    f"message is {expected_excel_upload_status_text}",
                                                            pass_=False, log=True, screenshot=True)
                            raise Exception("Unable to find status message while overriding PRISMA Excel File.")
        except Exception:
            raise Exception("Unable to Override PRISMA Excel file")

    def add_picos_details(self, locatorname, filepath, pop_data, stdy_data, env, project):
        expected_status_text = "Saved successfully"

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)
        expected_row_headers = self.exbase.get_individual_col_data(filepath, locatorname, 'Sheet1', 'Row_headers')
        expected_col_headers = self.exbase.get_individual_col_data(filepath, locatorname, 'Sheet1', 'Col_headers')

        try:
            for i in pop_data:
                for j in stdy_data:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']

                    protocol_picos_data = []

                    for k in stdytype:
                        if project == "Non-Oncology" and k == 'Clinical-Interventional':
                            data = self.exbase.get_individual_col_data(filepath, locatorname, 'Sheet1',
                                                                       'data_intervention')
                        elif project == "Non-Oncology" and k == 'Clinical-RWE':
                            data = self.exbase.get_individual_col_data(filepath, locatorname, 'Sheet1', 'data_rwe')
                        else:
                            data = self.exbase.get_individual_col_data(filepath, locatorname, 'Sheet1', 'data')

                        selected_pop_val = self.base.selectbyvisibletext("picos_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("picos_study_type_dropdown", k, env)
                        time.sleep(1)

                        actual_row_headers = self.get_texts("row_header_values", env)

                        actual_col_headers = self.get_texts("col_header_values", env)
                        # Removing the empty column name
                        actual_col_headers.pop(0)

                        row_header_comparison = self.slrreport.list_comparison_between_reports_data(
                            expected_row_headers, actual_row_headers)

                        if len(row_header_comparison) == 0:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PICOS page row headers are displayed as expected for Population -> {i[0]}, "
                                        f"Study Type -> {k}.", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Mismatch found in PICOS page row headers for Population -> {i[0]}, "
                                        f"Study Type -> {k}. Mismatch values are arranged in following order -> "
                                        f"Expected row headers, Actual row headers. {row_header_comparison}",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(f"Mismatch found in PICOS page row headers for Population -> {i[0]}, "
                                            f"Study Type -> {k}.")

                        col_header_comparison = self.slrreport.list_comparison_between_reports_data(
                            expected_col_headers, actual_col_headers)

                        if len(col_header_comparison) == 0:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PICOS page col headers are displayed as expected for Population -> {i[0]}, "
                                        f"Study Type -> {k}.", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Mismatch found in PICOS page col headers for Population -> {i[0]}, "
                                        f"Study Type -> {k}. Mismatch values are arranged in following order -> "
                                        f"Expected col headers, Actual col headers. {col_header_comparison}",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(f"Mismatch found in PICOS page row headers for Population -> {i[0]}, "
                                            f"Study Type -> {k}.")

                        # Enter values in PICOS page
                        picos_data = []
                        data_eles = self.select_elements('row_data', env)
                        idx = 0
                        for locator in data_eles:
                            if project == "Oncology":
                                if idx > 2:
                                    idx = 0
                            else:
                                if idx > 3:
                                    idx = 0
                            locator.clear()
                            locator.send_keys(f"{data[idx]}")
                            picos_data.append(data[idx].split('\n'))
                            idx += 1

                        # This will give one single list with all the picos data added
                        added_picos_data = list(np.concatenate(picos_data))

                        self.scrollback('picos_page_heading', env)
                        self.jsclick("picos_save_btn", env)
                        time.sleep(1)

                        actual_status_text = self.get_status_text("prisma_excel_status_text", env)

                        if actual_status_text == expected_status_text:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Addition of PICOS data is success for Population -> {i[0]}, "
                                        f"Study Type -> {k}.",
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Unable to find status message while adding PICOS data for Population -> "
                                        f"{i[0]}, Study Type -> {k}. Actual status message is {actual_status_text} "
                                        f"and Expected status message is {expected_status_text}",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while adding PICOS data for Population -> "
                                            f"{i[0]}, Study Type -> {k}.")

                        self.refreshpage()
                        time.sleep(2)
                        protocol_picos_data.append(added_picos_data)
                    return protocol_picos_data
        except Exception:
            raise Exception("Unable to add PICOS data")

    def validate_view_picos(self, locatorname, filepath, pop_data, stdy_data, env, project):
        try:
            for i in pop_data:
                for j in stdy_data:
                    self.go_to_nested_page("protocol_link", "picos", env)
                    pop = [[i[k]] for k in range(0, len(i))]
                    stdy = [[j[k]] for k in range(0, len(j))]
                    protocol_picos_data = self.add_picos_details(locatorname, filepath, pop_data, stdy_data, env, project)

                    self.base.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(i[0], i[1], env)
                    self.slrreport.select_data(j[0], j[1], env)

                    self.click("view_picos_button", env)
                    time.sleep(1)

                    if project == "Non-Oncology" and self.isdisplayed("view_picos_clinical_intervention_tab", env):
                        self.click("view_picos_clinical_intervention_tab", env)
                        view_picos_data = self.get_texts("view_picos_active_tab_data", env)
                        picos_data_comparison_intvn = self.slrreport.list_comparison_between_reports_data(
                            protocol_picos_data[0], view_picos_data)

                        if len(picos_data_comparison_intvn) == 0:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-Interventional PICOS data under Search LiveSLR -> "
                                        f"View PICOS section is matching with PICOS data under Protocol -> PICOS "
                                        f"and Inc-Exc criteria page for Population -> {i[0]}, Study Type -> {j[0]}.",
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-Interventional Mismatch found in PICOS data under "
                                        f"Search LiveSLR -> View PICOS section with PICOS data under Protocol -> "
                                        f"PICOS and Inc-Exc criteria page for Population -> {i[0]}, Study Type -> "
                                        f"{j[0]}. Mismatch values are arranged in following order -> Protocol PICOS "
                                        f"Data, View PICOS Data. {picos_data_comparison_intvn}",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"Non-Oncology Clinical-Interventional Mismatch found in PICOS data "
                                            f"under Search LiveSLR -> View PICOS section with PICOS data under "
                                            f"Protocol -> PICOS and Inc-Exc criteria page for Population -> "
                                            f"{i[0]}, Study Type -> {j[0]}.")
                    if project == "Non-Oncology" and self.isdisplayed("view_picos_clinical_rwe_tab", env):
                        self.click("view_picos_clinical_rwe_tab", env)
                        view_picos_data = self.get_texts("view_picos_active_tab_data", env)
                        picos_data_comparison_rwe = self.slrreport.list_comparison_between_reports_data(
                            protocol_picos_data[1], view_picos_data)

                        if len(picos_data_comparison_rwe) == 0:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-RWE PICOS data under Search LiveSLR -> View PICOS "
                                        f"section is matching with PICOS data under Protocol -> PICOS and Inc-Exc "
                                        f"criteria page for Population -> {i[0]}, Study Type -> {j[0]}.",
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-RWE Mismatch found in PICOS data under Search "
                                        f"LiveSLR -> View PICOS section with PICOS data under Protocol -> PICOS "
                                        f"and Inc-Exc criteria page for Population -> {i[0]}, Study Type -> {j[0]}. "
                                        f"Mismatch values are arranged in following order -> Protocol PICOS Data, "
                                        f"View PICOS Data. {picos_data_comparison_rwe}",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"Non-Oncology Clinical-RWE Mismatch found in PICOS data under Search "
                                            f"LiveSLR -> View PICOS section with PICOS data under Protocol -> "
                                            f"PICOS and Inc-Exc criteria page for Population -> "
                                            f"{i[0]}, Study Type -> {j[0]}.")
                    else:
                        view_picos_data = self.get_texts("view_picos_data", env)
                        picos_data_comparison = self.slrreport.list_comparison_between_reports_data(
                            protocol_picos_data[0], view_picos_data)

                        if len(picos_data_comparison) == 0:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PICOS data under Search LiveSLR -> View PICOS section is matching with "
                                        f"PICOS data under Protocol -> PICOS and Inc-Exc criteria page for "
                                        f"Population -> {i[0]}, Study Type -> {j[0]}.",
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Mismatch found in PICOS data under Search LiveSLR -> View PICOS section "
                                        f"with PICOS data under Protocol -> PICOS and Inc-Exc criteria page for "
                                        f"Population -> {i[0]}, Study Type -> {j[0]}. Mismatch values are arranged "
                                        f"in following order -> Protocol PICOS Data, View PICOS Data. "
                                        f"{picos_data_comparison}",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"Mismatch found in PICOS data under Search LiveSLR -> View PICOS "
                                            f"section with PICOS data under Protocol -> PICOS and Inc-Exc criteria "
                                            f"page for Population -> {i[0]}, Study Type -> {j[0]}.")
                    self.refreshpage()
        except Exception:
            raise Exception("Error in accessing View PICOS")

    def add_valid_search_strategy_details(self, locatorname, filepath, pop_data, stdy_data, env, project, fname):
        expected_excel_upload_status_text = "Search strategy updated successfully"

        today = date.today()
        # Manipulating the date values when day point to month end
        if today.day in [30, 31]:
            day_val = (today - timedelta(10)).strftime("%d")
        else:
            day_val = today.day

        df = pd.read_excel(filepath)

        self.refreshpage()
        time.sleep(2)
        try:
            for i in pop_data:
                for j in stdy_data:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    fileupload = data1_val["Files_to_upload"]
                    fileupload = [item for item in fileupload if str(item) != 'nan']
                    db_val = data1_val["db_search_val"]
                    db_val = [item for item in db_val if str(item) != 'nan']
                    template = data1_val["Template_name_strategy"]
                    template = [item for item in template if str(item) != 'nan']

                    upload_data = [[stdytype[i], fileupload[i], db_val[i], template[i]] for i in
                                   range(0, len(stdytype))]
                    table_data = []
                    entered_db_value = []

                    for k in upload_data:
                        selected_pop_val = self.base.selectbyvisibletext("searchstrategy_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("searchstrategy_study_type_dropdown",
                                                                         k[0], env)
                        time.sleep(1)

                        self.slrreport.generate_download_report("searchstrategy_template_download_btn", env)
                        # Renaming the filename because there is an issue in downloading filenames with same name
                        # multiple times in headless mode. So as an alternative renaming the file after downloading
                        # in each iteration
                        self.file_rename(self.slrreport.get_latest_filename(UnivWaitFor=180),
                                         f"{k[0]}_search-strategy-template_{project}_{fname}.xlsx")
                        expected_template_name = f"{k[3]}_{fname}.xlsx"
                        downloaded_template_name = self.slrreport.get_latest_filename(UnivWaitFor=180)
                        if downloaded_template_name == expected_template_name:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Correct Template is downloaded. Template name is {downloaded_template_name}",
                                pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Mismatch in search strategy template name. Expected Template name is "
                                        f"{expected_template_name} and Actual Template name is "
                                        f"{downloaded_template_name}",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(f"Mismatch in search strategy template name.")

                        self.click("searchstrategy_date", env)
                        self.select_calendar_date(day_val)

                        self.input_text("searchstrategy_dbsearch", k[2], env)

                        jscmd = ReadConfig.get_remove_att_JScommand(17, 'hidden')
                        self.jsclick_hide(jscmd)
                        self.input_text("searchstrategy_upload_file", os.getcwd()+"\\"+k[1], env)

                        self.jsclick("searchstrategy_upload_btn", env)
                        time.sleep(2)

                        actual_excel_upload_status_text = self.get_status_text("searchstrategy_status_text", env)
                        # time.sleep(2)

                        if actual_excel_upload_status_text == expected_excel_upload_status_text:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Addition of Search strategy data is success for Population -> {i[0]}, "
                                        f"Study Type -> {k[0]}.",
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Unable to find status message while adding Search strategy data for "
                                        f"Population -> {i[0]}, Study Type -> {k[0]}. Actual message is "
                                        f"{actual_excel_upload_status_text} and Expected message is "
                                        f"{expected_excel_upload_status_text}",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while adding Search strategy data for "
                                            f"Population -> {i[0]}, Study Type -> {k[0]}.")

                        uploaded_tabledata = self.export_web_table(
                            "table table-bordered table-striped table-sm search-terms",
                            f"uploadedsearchstrategydata_{k[0]}_{project}_{fname}")
                        self.refreshpage()
                        time.sleep(2)
                        table_data.append(uploaded_tabledata)
                        entered_db_value.append(k[2])
                    res = [[table_data[i], entered_db_value[i]] for i in range(0, len(table_data))]
                    return res
        except Exception:
            raise Exception("Unable to add Search strategy data")

    def add_invalid_search_strategy_details(self, locatorname, filepath, pop_data, stdy_data, env):
        expected_error_text = "Error while updating search strategy: The file extension should belong to this " \
                              "list: [.xls, .xlsx]"

        today = date.today()
        # Manipulating the date values when day point to month end
        if today.day in [30, 31]:
            day_val = (today - timedelta(10)).strftime("%d")
        else:
            day_val = today.day

        df = pd.read_excel(filepath)

        self.refreshpage()
        time.sleep(2)
        try:
            for i in pop_data:
                for j in stdy_data:
                    # Get StudyType and Files path to upload Managae QA Data
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']
                    fileupload = data1_val["Invalid_Files"]
                    fileupload = [item for item in fileupload if str(item) != 'nan']
                    db_val = data1_val["db_search_val"]
                    db_val = [item for item in db_val if str(item) != 'nan']
                    template = data1_val["Template_name_strategy"]
                    template = [item for item in template if str(item) != 'nan']

                    upload_data = [[stdytype[i], fileupload[i], db_val[i], template[i]] for i in
                                   range(0, len(stdytype))]

                    for k in upload_data:
                        selected_pop_val = self.base.selectbyvisibletext("searchstrategy_pop_dropdown", i[0], env)
                        time.sleep(1)

                        selected_slr_val = self.base.selectbyvisibletext("searchstrategy_study_type_dropdown",
                                                                         k[0], env)
                        self.LogScreenshot.fLogScreenshot(message=f"Selected Population and SLR Type Details: ",
                                                          pass_=True, log=True, screenshot=True)
                        time.sleep(1)

                        self.click("searchstrategy_date", env)
                        self.select_calendar_date(day_val)

                        self.input_text("searchstrategy_dbsearch", k[2], env)

                        jscmd = ReadConfig.get_remove_att_JScommand(17, 'hidden')
                        self.jsclick_hide(jscmd)
                        self.input_text("searchstrategy_upload_file", os.getcwd()+"\\"+k[1], env)

                        self.jsclick("searchstrategy_upload_btn", env)
                        time.sleep(2)

                        actual_error_text = self.get_status_text("searchstrategy_status_text", env)
                        # time.sleep(2)

                        if actual_error_text == expected_error_text:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"File with invalid format is not uploaded as expected. Invalid file is "
                                        f"'{os.path.basename(k[1])}'",
                                pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Unable to find status message while adding Search strategy data for "
                                        f"Population -> {i[0]}, Study Type -> {k[0]}. Actual status message is "
                                        f"{actual_error_text} and Expected status message is {expected_error_text}",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"Unable to find status message while adding Search strategy data for "
                                            f"Population -> {i[0]}, Study Type -> {k[0]}.")

                        self.refreshpage()
                        time.sleep(2)
        except Exception:
            raise Exception("Unable to add Search strategy data")

    def validate_view_search_strategy(self, locatorname, filepath, pop_data, stdy_data, env, prjname):
        try:
            for i in pop_data:
                for j in stdy_data:
                    self.go_to_nested_page("protocol_link", "searchstrategy", env)
                    pop = [[i[k]] for k in range(0, len(i))]
                    stdy = [[j[k]] for k in range(0, len(j))]
                    uploaded_data = self.add_valid_search_strategy_details(locatorname, filepath, pop_data, stdy_data,
                                                                           env, prjname, "viewstrategy")

                    self.base.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(i[0], i[1], env)
                    self.slrreport.select_data(j[0], j[1], env)

                    self.click("view_searchstrategy_button", env)
                    time.sleep(1)

                    if prjname == "Non-Oncology" and self.isdisplayed("view_search_clinical_intervention_tab", env):
                        self.click("view_search_clinical_intervention_tab", env)
                        view_date_val = self.get_text("view_search_date_active_tab_data", env)
                        database_search_val_intvn = self.get_text("view_search_database_active_tab_data", env)

                        nononco_view_strategy_filename_intvn = self.export_web_table(
                            "table table-bordered table-striped table-sm search-term",
                            f"viewsearchstrategydata_Clinical-Interventional_{prjname}")

                        nononco_searchstrategy_data_intvn = pd.read_excel(
                            f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}', usecols='B:D')
                        nononco_view_searchstrategy_data_intvn = pd.read_excel(
                            f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_intvn}', usecols='B:D')

                        nononco_view_searchstrategy_data_intvn.rename(columns={'Unnamed: 0': 'Line'}, inplace=True)

                        if nononco_searchstrategy_data_intvn.equals(nononco_view_searchstrategy_data_intvn) and \
                                uploaded_data[0][1] == database_search_val_intvn:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-Interventional File contents between Uploaded Search "
                                        f"Strategy File "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}').name}' "
                                        f"and View Search Strategy Data "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_intvn}').name}' " 
                                        f"are matching", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-Interventional Strategy DB value : "
                                        f"{uploaded_data[0][1]} and View DB value : {database_search_val_intvn}",
                                pass_=False, log=True, screenshot=True)
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-Interventional File contents between Uploaded Search "
                                        f"Strategy File "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}').name}' "
                                        f"and View Search Strategy Data "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_intvn}').name}' "
                                        f"are not matching", pass_=False, log=True, screenshot=True)
                            raise Exception(f"Non-Oncology Clinical-Interventional File contents between Uploaded "
                                            f"Search Strategy File "
                                            f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}').name}'"
                                            f"and View Search Strategy Data "
                                            f"'{Path(f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_intvn}').name}' "
                                            f"are not matching")
                    if prjname == "Non-Oncology" and self.isdisplayed("view_search_clinical_rwe_tab", env):
                        self.click("view_search_clinical_rwe_tab", env)
                        view_date_val = self.get_text("view_search_date_active_tab_data", env)
                        database_search_val_rwe = self.get_text("view_search_database_active_tab_data", env)

                        nononco_view_strategy_filename_rwe = self.export_web_table(
                            "table table-bordered table-striped table-sm search-term",
                            f"viewsearchstrategydata_Clinical-RWE_{prjname}")

                        nononco_searchstrategy_data_rwe = pd.read_excel(
                            f'ActualOutputs//web_table_exports//{uploaded_data[1][0]}', usecols='B:D')
                        nononco_view_searchstrategy_data_rwe = pd.read_excel(
                            f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_rwe}', usecols='B:D')

                        nononco_view_searchstrategy_data_rwe.rename(columns={'Unnamed: 0': 'Line'}, inplace=True)

                        if nononco_searchstrategy_data_rwe.equals(nononco_view_searchstrategy_data_rwe) and \
                                uploaded_data[1][1] == database_search_val_rwe:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-RWE File contents between Uploaded Search "
                                        f"Strategy File "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[1][0]}').name}' "
                                        f"and View Search Strategy Data "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_rwe}').name}' "
                                        f"are matching", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-RWE Strategy DB value : {uploaded_data[1][1]} "
                                        f"and View DB value : {database_search_val_rwe}",
                                pass_=False, log=True, screenshot=True)
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Non-Oncology Clinical-RWE File contents between Uploaded Search Strategy File"
                                        f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[1][0]}').name}' "
                                        f"and View Search Strategy Data "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_rwe}').name}' "
                                        f"are not matching", pass_=False, log=True, screenshot=True)
                            raise Exception(f"File contents between Uploaded Search Strategy File "
                                            f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[1][0]}').name}'"
                                            f"and View Search Strategy Data "
                                            f"'{Path(f'ActualOutputs//web_table_exports//{nononco_view_strategy_filename_rwe}').name}' "
                                            f"are not matching")
                    else:
                        view_date_val = self.get_text("view_search_date", env)
                        database_search_val = self.get_text("view_search_database", env)

                        onco_view_strategy_filename = self.export_web_table(
                            "table table-bordered table-striped table-sm search-term",
                            f"viewsearchstrategydata_{j[0]}_{prjname}")

                        onco_searchstrategy_data = pd.read_excel(
                            f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}', usecols='B:D')
                        onco_view_searchstrategy_data = pd.read_excel(
                            f'ActualOutputs//web_table_exports//{onco_view_strategy_filename}', usecols='B:D')

                        onco_view_searchstrategy_data.rename(columns={'Unnamed: 0': 'Line'}, inplace=True)

                        if onco_searchstrategy_data.equals(onco_view_searchstrategy_data) and \
                                uploaded_data[0][1] == database_search_val:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"File contents between Uploaded Search Strategy File "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}').name}' "
                                        f"and View Search Strategy Data "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{onco_view_strategy_filename}').name}' "
                                        f"are matching", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"Strategy DB value : {uploaded_data[0][1]} and View DB value : "
                                        f"{database_search_val}",
                                pass_=False, log=True, screenshot=True)
                            self.LogScreenshot.fLogScreenshot(
                                message=f"File contents between Uploaded Search Strategy File "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}').name}' "
                                        f"and View Search Strategy Data "
                                        f"'{Path(f'ActualOutputs//web_table_exports//{onco_view_strategy_filename}').name}' "
                                        f"are not matching", pass_=False, log=True, screenshot=True)
                            raise Exception(f"File contents between Uploaded Search Strategy File "
                                            f"'{Path(f'ActualOutputs//web_table_exports//{uploaded_data[0][0]}').name}'"
                                            f"and View Search Strategy Data "
                                            f"'{Path(f'ActualOutputs//web_table_exports//{onco_view_strategy_filename}').name}' "
                                            f"are not matching")
                    self.refreshpage()
        except Exception:
            raise Exception("Unable to add Search strategy data")

    def download_protocol_file(self, locatorname, filepath, index, pop_data, stdy_data, env, add_criteria, project):
        expected_data_filepath = self.exbase.get_template_file_details(filepath, locatorname,
                                                                       "ExpectedSourceTemplateFile")

        try:
            for i in pop_data:
                for j in stdy_data:
                    self.go_to_page("SLR_Dashboard", env)

                    self.go_to_nested_page("protocol_link", "picos", env)
                    picos_data = self.add_picos_details(locatorname, filepath, pop_data, stdy_data, env, project)

                    self.go_to_page("searchstrategy", env)
                    uploaded_data = self.add_valid_search_strategy_details(locatorname, filepath, pop_data, stdy_data, env,
                                                                           project, f"downloadprotocol_{index}")

                    # Go to live slr page
                    self.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(f"{i[0]}", f"{i[1]}", env)
                    self.slrreport.select_data(f"{j[0]}", f"{j[1]}", env)
                    if len(add_criteria) != 0:
                        for v in add_criteria:
                            self.slrreport.select_sub_section(f"{v[0]}", f"{v[1]}", env, f"{v[2]}")

                    time.sleep(1)
                    self.slrreport.generate_download_report("Download_protocol_btn", env)
                    protocol_filename = self.slrreport.get_and_validate_filename(filepath)

                    self.slrreport.generate_download_report("excel_report", env)
                    excel_filename = self.slrreport.get_and_validate_filename(filepath)

                    expected_workbook = openpyxl.load_workbook(f'{expected_data_filepath}')

                    if project == 'Oncology':
                        expected_sheets = ['PICOS', 'INC-EXC', 'Search Strategies']
                    elif project == 'Non-Oncology':
                        expected_sheets = ['PICOS', 'Inclusion Exclusion Criteria', 'Search Strategies']

                    for sheet in expected_sheets:
                        if sheet in expected_workbook.sheetnames:
                            expecteddata = pd.read_excel(f'{expected_data_filepath}', sheet_name=sheet)
                            actualdata_protocolfile = pd.read_excel(f'ActualOutputs//{protocol_filename}', sheet_name=sheet)
                            if project == "Non-Oncology" and sheet == "Search Strategies":
                                sheet = "Search Strategy(ies)"
                            elif sheet == "Search Strategies":
                                sheet = "SEARCH STRATEGIES"
                            actualdata_compexcelfile = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name=sheet)

                            # Removing the 'Back To Toc' column from complete excel report to compare the exact data
                            # with expected file
                            actualdata_compexcelfile = actualdata_compexcelfile.iloc[:, :-1]

                            if expecteddata.equals(actualdata_protocolfile):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel "
                                            f"Report '{Path(f'ActualOutputs//{protocol_filename}').name}' are matching",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel "
                                            f"Report '{Path(f'ActualOutputs//{protocol_filename}').name}' are not matching",
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(
                                    f"From '{sheet}' sheet -> File contents between Expected File "
                                    f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel Report "
                                    f"'{Path(f'ActualOutputs//{protocol_filename}').name}' are not matching")

                            '''Renaming the Column name as there is slight difference in Column header in Complete 
                            Excel report for INC-EXC and SEARCH STRATEGIES sheets'''
                            if project == "Oncology":
                                if sheet == "INC-EXC":
                                    expecteddata.rename(columns={'INCLUSION AND EXCLUSION CRITERIA: Test_Automation_1': f'INCLUSION AND EXCLUSION CRITERIA: Test_Automation_1 {j[0]}'}, inplace=True)
                                if sheet == "SEARCH STRATEGIES":
                                    expecteddata.rename(columns={'SEARCH STRATEGY: Test_Automation_1': f'SEARCH STRATEGY: Test_Automation_1 {j[0]}'}, inplace=True)

                            if project == "Non-Oncology":
                                if sheet == "PICOS":
                                    expecteddata.rename(columns={'PICOS: Test_NonOncology_Automation_3': 'PICOS: Test_NonOncology_Automation_3 Clinical'}, inplace=True)
                                    expecteddata.rename(columns={'PICOS: Test_NonOncology_Automation_3.1': 'PICOS: Test_NonOncology_Automation_3 Clinical.1'}, inplace=True)
                                if sheet == "Inclusion Exclusion Criteria":
                                    expecteddata.rename(columns={'Inclusion and Exclusion Criteria: Test_NonOncology_Automation_3': 'Inclusion and Exclusion Criteria: Test_NonOncology_Automation_3 Clinical'}, inplace=True)
                                if sheet == "Search Strategy(ies)":
                                    expecteddata.rename(columns={'Search Strategy(ies): Test_NonOncology_Automation_3': 'Search Strategy(ies): Test_NonOncology_Automation_3 Clinical'}, inplace=True)

                            if expecteddata.equals(actualdata_compexcelfile):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel "
                                            f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are matching",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel "
                                            f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are not matching",
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(
                                    f"From '{sheet}' sheet -> File contents between Expected File "
                                    f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel Report "
                                    f"'{Path(f'ActualOutputs//{excel_filename}').name}' are not matching")

        except Exception:
            raise Exception("Error while validating Protocol Excel File")

    def validate_viewprisma_and_prisma_protocol_file_withoutdata(self, locatorname, filepath, pop_data, stdy_data, env, add_criteria, project):
        expected_data_filepath = self.exbase.get_template_file_details(filepath, locatorname,
                                                                       "ExpectedSourceTemplateFile_withoutdata")

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data:
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']

                    self.go_to_nested_page("protocol_link", "prismas", env)
                    self.del_prisma_excel(locatorname, filepath, pop_data, stdy_data, env)

                    self.base.go_to_page("SLR_Homepage", env)
                    self.slrreport.select_data(i[0], i[1], env)
                    self.slrreport.select_data(j[0], j[1], env)
                    if len(add_criteria) != 0:
                        for v in add_criteria:
                            self.slrreport.select_sub_section(f"{v[0]}", f"{v[1]}", env, f"{v[2]}")

                    '''Validating the absence of data in PRISMA Protocol file and Complete Excel Report when Excel file is deleted in Protocol->PRISMA page'''
                    time.sleep(1)
                    if self.clickable('Download_protocol_btn', env):
                        self.LogScreenshot.fLogScreenshot(message=f"Download Protocol button is enabled after selecting the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=True, log=True, screenshot=True)
                        self.slrreport.generate_download_report("Download_protocol_btn", env)
                        protocol_filename = self.slrreport.get_and_validate_filename(filepath)

                        self.slrreport.generate_download_report("excel_report", env)
                        excel_filename = self.slrreport.get_and_validate_filename(filepath)
                    else:
                        self.LogScreenshot.fLogScreenshot(message=f"Download Protocol button is not enabled after selecting the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=True, log=True, screenshot=True)
                        raise Exception(f"Download Protocol button is not enabled after selecting the Population '{i[0]}' -> SLR Type '{j[0]}'")

                    expected_workbook = openpyxl.load_workbook(f'{expected_data_filepath}')

                    for sheet in expected_workbook.sheetnames:
                        expecteddata = pd.read_excel(f'{expected_data_filepath}', sheet_name=sheet)
                        actualdata_protocolfile = pd.read_excel(f'ActualOutputs//{protocol_filename}', sheet_name=sheet)
                        actualdata_compexcelfile = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name=sheet)

                        # Removing the 'Back To Toc' column from complete excel report to compare the exact data
                        # with expected file
                        actualdata_compexcelfile = actualdata_compexcelfile.iloc[:, 1:]
                        actualdata_compexcelfile.rename(columns={'Unnamed: 1': 'Unnamed: 0'}, inplace=True)

                        if expecteddata.equals(actualdata_protocolfile):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"From '{sheet}' sheet -> File contents between Expected File "
                                        f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel "
                                        f"Report '{Path(f'ActualOutputs//{protocol_filename}').name}' are matching when PRISMA Excel file is deleted in Protocol->PRISMA page",
                                pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"From '{sheet}' sheet -> File contents between Expected File "
                                        f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel "
                                        f"Report '{Path(f'ActualOutputs//{protocol_filename}').name}' are not matching when PRISMA Excel file is deleted in Protocol->PRISMA page",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(
                                f"From '{sheet}' sheet -> File contents between Expected File "
                                f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel Report "
                                f"'{Path(f'ActualOutputs//{protocol_filename}').name}' are not matching when PRISMA Excel file is deleted in Protocol->PRISMA page")
                        
                        if expecteddata.equals(actualdata_compexcelfile):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"From '{sheet}' sheet -> File contents between Expected File "
                                        f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel "
                                        f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are matching when PRISMA Excel file is deleted in Protocol->PRISMA page",
                                pass_=True, log=True, screenshot=False)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"From '{sheet}' sheet -> File contents between Expected File "
                                        f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel "
                                        f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are not matching when PRISMA Excel file is deleted in Protocol->PRISMA page",
                                pass_=False, log=True, screenshot=False)
                            raise Exception(
                                f"From '{sheet}' sheet -> File contents between Expected File "
                                f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel Report "
                                f"'{Path(f'ActualOutputs//{excel_filename}').name}' are not matching when PRISMA Excel file is deleted in Protocol->PRISMA page")

                    '''Validating the absence of data in View PRISMA when Excel file is deleted in Protocol->PRISMA page'''
                    self.click("view_prisma_button", env)
                    time.sleep(1)

                    if project == "Non-Oncology" and self.isdisplayed("view_prisma_clinical_intervention_tab", env) and 'Clinical-Interventional' in stdytype:
                        self.click("view_prisma_clinical_intervention_tab", env)
                        if not self.isdisplayed("view_prisma_clinical_intvtn_tab_data", env):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
                    time.sleep(2)
                    if project == "Non-Oncology" and self.isdisplayed("view_prisma_clinical_rwe_tab", env) and 'Clinical-RWE' in stdytype:
                        self.click("view_prisma_clinical_rwe_tab", env)
                        if not self.isdisplayed("view_prisma_clinical_rwe_tab_data", env):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
                    if project == "Oncology": 
                        if not self.isdisplayed("view_prisma_data", env):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")

                    # Closing the View Prisma window
                    self.click("view_prisma_close_button", env)
                    time.sleep(2)

                    # Checking the presence of View Updated Prisma button when additional criteria has been selected
                    if self.isdisplayed("view_updated_prisma_btn", env) and self.clickable("view_updated_prisma_btn", env):
                        self.LogScreenshot.fLogScreenshot(message=f"View Updated Prisma button is displayed and enabled after selecting the additional criteria options for the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=True, log=True, screenshot=True)
                        self.click("view_updated_prisma_btn", env)

                        if project == "Non-Oncology" and self.isdisplayed("view_updated_prisma_clinical_intervention_tab", env) and 'Clinical-Interventional' in stdytype:
                            self.click("view_updated_prisma_clinical_intervention_tab", env)
                            if not self.isdisplayed("view_updated_prisma_clinical_intvtn_tab_data", env):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is not present in View Updated PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
                        if project == "Non-Oncology" and self.isdisplayed("view_updated_prisma_clinical_rwe_tab", env) and 'Clinical-RWE' in stdytype:
                            self.click("view_updated_prisma_clinical_rwe_tab", env)
                            if not self.isdisplayed("view_updated_prisma_clinical_rwe_tab_data", env):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is not present in View Updated PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
                        if project == "Oncology":
                            if not self.isdisplayed("view_updated_prisma_data", env):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Updated PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")                        
                    
                        # Closing the View Updated Prisma window
                        self.click("view_updated_prisma_close_btn", env)
                    elif self.isdisplayed("view_updated_prisma_btn", env) and not self.clickable("view_updated_prisma_btn", env):
                        self.LogScreenshot.fLogScreenshot(message=f"View Updated Prisma button is displayed and not enabled after selecting the additional criteria options for the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=False, log=True, screenshot=True)
                        raise Exception(f"View Updated Prisma button is displayed and not enabled after selecting the additional criteria options for the Population '{i[0]}' -> SLR Type '{j[0]}'")

                    self.refreshpage()
        except Exception:
            raise Exception("Unable to View Original PRISMA data")

    def validate_viewprisma_and_prisma_protocol_file(self, locatorname, filepath, index, pop_data, stdy_data, env, add_criteria, project):
        expected_data_filepath = self.exbase.get_template_file_details(filepath, locatorname,
                                                                       "ExpectedSourceTemplateFile")

        # This Dataframe will be used to read the study type and study files based on the given SLR Type
        df = pd.read_excel(filepath)

        try:
            for i in pop_data:
                for j in stdy_data:
                    data1 = df[df["Name"] == locatorname]
                    data1_val = data1[data1["slrtype"] == j[0]]
                    stdytype = data1_val["Study_Types"]
                    stdytype = [item for item in stdytype if str(item) != 'nan']

                    self.go_to_nested_page("protocol_link", "prismas", env)
                    self.upload_prisma_excel(locatorname, filepath, pop_data, stdy_data, env, project, f"viewprisma_{index}")

                    self.base.go_to_page("SLR_Homepage", env)
                    self.refreshpage()
                    self.slrreport.select_data(i[0], i[1], env)
                    self.slrreport.select_data(j[0], j[1], env)
                    if len(add_criteria) != 0:
                        for v in add_criteria:
                            self.slrreport.select_sub_section(f"{v[0]}", f"{v[1]}", env, f"{v[2]}")

                    '''Validating the PRISMA Protocol file and Complete Excel Report data'''
                    time.sleep(1)
                    if self.clickable('Download_protocol_btn', env):
                        self.LogScreenshot.fLogScreenshot(message=f"Download Protocol button is enabled after selecting the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=True, log=True, screenshot=True)
                        self.slrreport.generate_download_report("Download_protocol_btn", env)
                        protocol_filename = self.slrreport.get_and_validate_filename(filepath)

                        self.slrreport.generate_download_report("excel_report", env)
                        excel_filename = self.slrreport.get_and_validate_filename(filepath)
                    else:
                        self.LogScreenshot.fLogScreenshot(message=f"Download Protocol button is not enabled after selecting the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=True, log=True, screenshot=True)
                        raise Exception(f"Download Protocol button is not enabled after selecting the Population '{i[0]}' -> SLR Type '{j[0]}'")

                    expected_workbook = openpyxl.load_workbook(f'{expected_data_filepath}')

                    if project == 'Oncology':
                        expected_sheets = ['PRISMA']
                    elif project == 'Non-Oncology':
                        expected_sheets = ['PRISMA-Intervtnl', 'PRISMA-RWE']

                    for sheet in expected_sheets:
                        if sheet in expected_workbook.sheetnames:
                            expecteddata = pd.read_excel(f'{expected_data_filepath}', sheet_name=sheet, skiprows=1)
                            actualdata_protocolfile = pd.read_excel(f'ActualOutputs//{protocol_filename}', sheet_name=sheet, skiprows=1)
                            actualdata_compexcelfile = pd.read_excel(f'ActualOutputs//{excel_filename}', sheet_name=sheet, skiprows=1)

                            if expecteddata.equals(actualdata_protocolfile):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel "
                                            f"Report '{Path(f'ActualOutputs//{protocol_filename}').name}' are matching",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel "
                                            f"Report '{Path(f'ActualOutputs//{protocol_filename}').name}' are not matching",
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(
                                    f"From '{sheet}' sheet -> File contents between Expected File "
                                    f"'{Path(f'{expected_data_filepath}').name}' and Protocol Excel Report "
                                    f"'{Path(f'ActualOutputs//{protocol_filename}').name}' are not matching")
                            
                            if expecteddata.equals(actualdata_compexcelfile):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel "
                                            f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are matching",
                                    pass_=True, log=True, screenshot=False)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"From '{sheet}' sheet -> File contents between Expected File "
                                            f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel "
                                            f"Report '{Path(f'ActualOutputs//{excel_filename}').name}' are not matching",
                                    pass_=False, log=True, screenshot=False)
                                raise Exception(
                                    f"From '{sheet}' sheet -> File contents between Expected File "
                                    f"'{Path(f'{expected_data_filepath}').name}' and Complete Excel Report "
                                    f"'{Path(f'ActualOutputs//{excel_filename}').name}' are not matching")
                    
                    '''Validationg the View PRISMA functionality'''
                    self.click("view_prisma_button", env)
                    time.sleep(1)

                    if project == "Non-Oncology" and self.isdisplayed("view_prisma_clinical_intervention_tab", env) and 'Clinical-Interventional' in stdytype:
                        self.click("view_prisma_clinical_intervention_tab", env)
                        if self.isdisplayed("view_prisma_clinical_intvtn_tab_data", env):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is present in View Original PRISMA section", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is not present in View Original PRISMA section",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is not present in View Original PRISMA section")
                    time.sleep(2)
                    if project == "Non-Oncology" and self.isdisplayed("view_prisma_clinical_rwe_tab", env) and 'Clinical-RWE' in stdytype:
                        self.click("view_prisma_clinical_rwe_tab", env)
                        if self.isdisplayed("view_prisma_clinical_rwe_tab_data", env):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is present in View Original PRISMA section", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is not present in View Original PRISMA section",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is not present in View Original PRISMA section")
                    if project == "Oncology":
                        if self.isdisplayed("view_prisma_data", env):
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section", pass_=True, log=True, screenshot=True)
                        else:
                            self.LogScreenshot.fLogScreenshot(
                                message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section",
                                pass_=False, log=True, screenshot=True)
                            raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section")

                    # Closing the View Prisma window
                    self.click("view_prisma_close_button", env)
                    time.sleep(2)

                    # Checking the presence of View Updated Prisma button when additional criteria has been selected
                    if self.isdisplayed("view_updated_prisma_btn", env) and self.clickable("view_updated_prisma_btn", env):
                        self.LogScreenshot.fLogScreenshot(message=f"View Updated Prisma button is displayed and enabled after selecting the additional criteria options for the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=True, log=True, screenshot=True)
                        self.click("view_updated_prisma_btn", env)

                        if project == "Non-Oncology" and self.isdisplayed("view_updated_prisma_clinical_intervention_tab", env) and 'Clinical-Interventional' in stdytype:
                            self.click("view_updated_prisma_clinical_intervention_tab", env)
                            if self.isdisplayed("view_updated_prisma_clinical_intvtn_tab_data", env):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is present in View Updated PRISMA section", pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is not present in View Updated PRISMA section",
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-Interventional' is not present in View Updated PRISMA section")
                        if project == "Non-Oncology" and self.isdisplayed("view_updated_prisma_clinical_rwe_tab", env) and 'Clinical-RWE' in stdytype:
                            self.click("view_updated_prisma_clinical_rwe_tab", env)
                            if self.isdisplayed("view_updated_prisma_clinical_rwe_tab_data", env):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is present in View Updated PRISMA section", pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is not present in View Updated PRISMA section",
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type 'Clinical-RWE' is not present in View Updated PRISMA section")
                        if project == "Oncology":
                            if self.isdisplayed("view_updated_prisma_data", env):
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section", pass_=True, log=True, screenshot=True)
                            else:
                                self.LogScreenshot.fLogScreenshot(
                                    message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Updated PRISMA section",
                                    pass_=False, log=True, screenshot=True)
                                raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Updated PRISMA section")                        
                    
                        # Closing the View Updated Prisma window
                        self.click("view_updated_prisma_close_btn", env)
                    elif self.isdisplayed("view_updated_prisma_btn", env) and not self.clickable("view_updated_prisma_btn", env):
                        self.LogScreenshot.fLogScreenshot(message=f"View Updated Prisma button is displayed and not enabled after selecting the additional criteria options for the Population '{i[0]}' -> SLR Type '{j[0]}'",
                                                        pass_=False, log=True, screenshot=True)
                        raise Exception(f"View Updated Prisma button is displayed and not enabled after selecting the additional criteria options for the Population '{i[0]}' -> SLR Type '{j[0]}'")

                    self.refreshpage()
        except Exception:
            raise Exception("Unable to View Original PRISMA data")

    # def validate_viewprisma_without_data(self, locatorname, filepath, pop_data, stdy_data, env, add_criteria, prjname):

    #     try:
    #         for i in pop_data:
    #             for j in stdy_data:
    #                 self.go_to_nested_page("protocol_link", "prismas", env)
    #                 self.del_prisma_excel(locatorname, filepath, pop_data, stdy_data, env)

    #                 self.base.go_to_page("SLR_Homepage", env)
    #                 self.refreshpage()
    #                 self.slrreport.select_data(i[0], i[1], env)
    #                 self.slrreport.select_data(j[0], j[1], env)
    #                 if len(add_criteria) != 0:
    #                     for v in add_criteria:
    #                         self.slrreport.select_sub_section(f"{v[0]}", f"{v[1]}", env, f"{v[2]}")
                    
    #                 self.click("view_prisma_button", env)
    #                 time.sleep(1)

    #                 if prjname == "Non-Oncology" and self.isdisplayed("view_prisma_clinical_intervention_tab", env):
    #                     self.click("view_prisma_clinical_intervention_tab", env)
    #                     if not self.isdisplayed("view_prisma_clinical_intvtn_tab_data", env):
    #                         self.LogScreenshot.fLogScreenshot(
    #                             message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
    #                     else:
    #                         self.LogScreenshot.fLogScreenshot(
    #                             message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
    #                             pass_=False, log=True, screenshot=True)
    #                         raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
    #                 time.sleep(2)
    #                 if prjname == "Non-Oncology" and self.isdisplayed("view_prisma_clinical_rwe_tab", env):
    #                     self.click("view_prisma_clinical_rwe_tab", env)
    #                     if not self.isdisplayed("view_prisma_clinical_rwe_tab_data", env):
    #                         self.LogScreenshot.fLogScreenshot(
    #                             message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
    #                     else:
    #                         self.LogScreenshot.fLogScreenshot(
    #                             message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
    #                             pass_=False, log=True, screenshot=True)
    #                         raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
    #                 else:
    #                     if not self.isdisplayed("view_prisma_data", env):
    #                         self.LogScreenshot.fLogScreenshot(
    #                             message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
    #                     else:
    #                         self.LogScreenshot.fLogScreenshot(
    #                             message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
    #                             pass_=False, log=True, screenshot=True)
    #                         raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Original PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")

    #                 # Closing the View Prisma window
    #                 self.click("view_prisma_close_button", env)
    #                 time.sleep(2)

    #                 # Checking the presence of View Updated Prisma button when additional criteria has been selected
    #                 if self.isdisplayed("view_updated_prisma_btn", env):
    #                     self.click("view_updated_prisma_btn", env)

    #                     if prjname == "Non-Oncology" and self.isdisplayed("view_updated_prisma_clinical_intervention_tab", env):
    #                         self.click("view_updated_prisma_clinical_intervention_tab", env)
    #                         if not self.isdisplayed("view_updated_prisma_clinical_intvtn_tab_data", env):
    #                             self.LogScreenshot.fLogScreenshot(
    #                                 message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Updated PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
    #                         else:
    #                             self.LogScreenshot.fLogScreenshot(
    #                                 message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
    #                                 pass_=False, log=True, screenshot=True)
    #                             raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
    #                     if prjname == "Non-Oncology" and self.isdisplayed("view_updated_prisma_clinical_rwe_tab", env):
    #                         self.click("view_updated_prisma_clinical_rwe_tab", env)
    #                         if not self.isdisplayed("view_updated_prisma_clinical_rwe_tab_data", env):
    #                             self.LogScreenshot.fLogScreenshot(
    #                                 message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Updated PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
    #                         else:
    #                             self.LogScreenshot.fLogScreenshot(
    #                                 message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
    #                                 pass_=False, log=True, screenshot=True)
    #                             raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")
    #                     else:                    
    #                         if not self.isdisplayed("view_updated_prisma_data", env):
    #                             self.LogScreenshot.fLogScreenshot(
    #                                 message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is not present in View Original PRISMA section as expected because the PRISMA excel file has been removed from Protocol -> PRISMA page", pass_=True, log=True, screenshot=True)
    #                         else:
    #                             self.LogScreenshot.fLogScreenshot(
    #                                 message=f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.",
    #                                 pass_=False, log=True, screenshot=True)
    #                             raise Exception(f"PRISMA Image for Population '{i[0]}' -> SLR Type '{j[0]}' is present in View Updated PRISMA section though PRISMA excel file is removed from Protocol -> PRISMA page which is not expected.")                        
                    
    #                     # Closing the View Updated Prisma window
    #                     self.click("view_updated_prisma_close_btn", env)

    #                 self.refreshpage()
    #     except Exception:
    #         raise Exception("Unable to View Original PRISMA data")
